from __future__ import annotations

import asyncio
import copy
import heapq
import io
import json
import logging
import math
import os
import statistics
import re
import shutil
import tempfile
import uuid
from collections import defaultdict, deque
from typing import Any, Dict, List, Optional, Tuple

import ezdxf
import pandas as pd
from fastapi import BackgroundTasks, Body, FastAPI, File, Form, Header, HTTPException, Query, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
from starlette.concurrency import run_in_threadpool

from .meissa_api import (
    Meissa2FARequired,
    meissa_get_carta_orthophoto_preview_png,
    meissa_orthophoto_disk_cache_best_valid_path_up_to,
    meissa_orthophoto_disk_cache_full_export_path_if_valid,
    meissa_orthophoto_disk_cache_path_if_valid,
    meissa_orthophoto_effective_preview_edge,
    meissa_orthophoto_write_disk_cache,
    meissa_orthophoto_write_disk_cache_full_export,
    meissa_orthophoto_full_export_crop_to_png_bytes,
    meissa_get_snapshot_overlay_2d_binary,
    meissa_get_snapshot_overlay_2d_image,
    meissa_get_snapshot_overlay_2d_georef,
    meissa_list_projects,
    meissa_sample_resource_points,
    meissa_list_snapshot_resources,
    meissa_list_snapshots,
    meissa_list_zones,
    meissa_login,
    meissa_login_with_verification,
    meissa_nearest_z_xy_combined,
    meissa_dsm_z_batch_from_carta_export,
)
from .construction_reports import (
    build_dashboard as build_construction_dashboard,
    delete_dataset as delete_construction_dataset,
    import_workbook_bytes as import_construction_workbook_bytes,
    list_datasets as list_construction_datasets,
    sync_pdam_workbook,
)
from .dxf_parser import foundation_pf_only_flag, parse_dxf_entities
from .excel_compare import compare_excel_workbook, inspect_excel_workbook
from .models import (
    ApplyFilterRequest,
    BuildingApplyRequest,
    BuildingDefinition,
    BuildingVertex,
    ConstructionDashboardRequest,
    MeissaLoginRequest,
    ConstructionSyncRequest,
    ExportFromClientBody,
    ManualHistoryMatchRefreshRequest,
    ManualMatchRequest,
    RecomputeRequest,
    BuildingSummary,
    CircleRecord,
    CircleResponse,
    DuplicateDetail,
    DuplicateGroup,
    FilterSettings,
    MatchError,
    MatchCorrection,
    MatchedText,
    SummaryStats,
    TextRecord,
)
from .pile_clustering import (
    ClusterConfig,
    PileCluster,
    PilePoint,
    cluster_piles,
)
from .pile_dataset_api import router as pile_dataset_router
from .work_context import (
    DEFAULT_PROJECT_NAME,
    DEFAULT_SOURCE_TYPE,
    apply_manual_history,
    apply_setting_defaults,
    apply_work_defaults,
    is_placeholder_building_name,
    normalize_project_name,
    normalize_source_type,
    update_manual_history,
)

app = FastAPI(
    title="DXF Circle/Text Service",
    version="0.3.1",
    description="Parses DXF files and exposes filtered circle/text data for the frontend viewer.",
)

logger = logging.getLogger(__name__)

# 직경 기본값: 도면 단위(m 등)에 맞춰 원형 폴리라인(예: Φ0.6)도 포함되도록 설정. 면적 필터와 연동.
DEFAULT_MIN_DIAMETER = 0.5
DEFAULT_MAX_DIAMETER = 0.65
DEFAULT_TEXT_HEIGHT_MIN = 0.4
DEFAULT_TEXT_HEIGHT_MAX = 1.1
DEFAULT_MAX_MATCH_DISTANCE = 2.0
COORD_PRECISION = 6
GRID_CELL_SIZE = 400.0
MAX_SEARCH_RADIUS = 3
# 긴 매칭(동료 대비 이상치) 재시도 시 격자 확장 — 인접 셀 텍스트가 첫 패스에서 누락된 경우
EXPANDED_MATCH_GRID_RADIUS = max(MAX_SEARCH_RADIUS * 2, 8)
MATCH_OUTLIER_REFINE_MAX_PASSES = 3
# 재매칭 트리거: 매칭 거리가 "파일(원) 직경" 대비 크고, 같은 컴포넌트 동료보다 유독 길 때
MATCH_OUTLIER_TO_DIAMETER_RATIO = 2.0
MATCH_OUTLIER_PEER_DISTANCE_RATIO = 1.5
MATCH_COST_SCALE = 1_000_000
# 말뚝번호 방향 힌트: 고정 각도가 아니라 (1) 수동 매칭 벡터 (2) 같은 유량 컴포넌트 안에서
# 각 원의 가장 가까운 텍스트 후보 방향을 가중 평균(가까울수록 가중)해 대세 축을 추정.
# 거리 항(MATCH_COST_SCALE)이 훨씬 커서 보통은 거리 우선. 0이면 방향 페널티 비활성.
LABEL_DIRECTION_ALIGN_COST_MAX = 8_000
# 수동 매칭 1건이 “가까운 후보 1표”에 상응하는 표 수(대세를 수동에 맞추기)
LABEL_DIRECTION_MANUAL_VOTE_EQUIV = 3.0
# 수동 잠금(둘 다 manualOverrides에 있음)이 서로 가까운 두 원에서 번호만 교환해도 거리 합이 확 줄면 제거
MANUAL_PRUNE_MAX_CENTER_FACTOR = 1.25
MANUAL_PRUNE_MIN_DISTANCE_IMPROVEMENT = 0.04
NUMERIC_LABEL_PATTERN = re.compile(r"^\d+$")
DEFAULT_CLUSTER_CONFIG = ClusterConfig()
PILE_NO_HEADER = "말뚝 NO."
PILE_X_HEADER = "X좌표(도면)"
PILE_Y_HEADER = "Y좌표(도면)"

# 프로젝트 루트: 실행 경로(cwd)와 무관하게 main.py 위치 기준으로 고정 (저장 데이터 경로 일관성)
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# 중앙 저장: DXF 처리 결과 저장 디렉터리
SAVED_WORKS_DIR = os.path.join(_PROJECT_ROOT, "data", "saved_works")
SAVED_WORKS_INDEX = "index.json"


def _ensure_saved_works_dir() -> str:
    os.makedirs(SAVED_WORKS_DIR, exist_ok=True)
    return SAVED_WORKS_DIR


def _read_saved_works_index() -> List[Dict[str, Any]]:
    _ensure_saved_works_dir()
    path = os.path.join(SAVED_WORKS_DIR, SAVED_WORKS_INDEX)
    if not os.path.isfile(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        items = data if isinstance(data, list) else []
        items = [apply_work_defaults(entry) for entry in items]
        # 기존 항목에 project 필드 없으면 기본값
        for entry in items:
            if "project" not in entry:
                entry["project"] = "기본"
        return items
    except Exception:
        return []


def _write_saved_works_index(items: List[Dict[str, Any]]) -> None:
    _ensure_saved_works_dir()
    path = os.path.join(SAVED_WORKS_DIR, SAVED_WORKS_INDEX)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)


def _work_file_path(work_id: str) -> str:
    safe_id = re.sub(r"[^\w\-]", "", work_id)
    return os.path.join(SAVED_WORKS_DIR, f"{safe_id}.json")


def _get_saved_work(work_id: str) -> Optional[Dict[str, Any]]:
    path = _work_file_path(work_id)
    if not os.path.isfile(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            payload = json.load(f)
        return apply_work_defaults(payload if isinstance(payload, dict) else {})
    except Exception:
        return None


def _put_saved_work(work_id: str, data: Dict[str, Any]) -> None:
    _ensure_saved_works_dir()
    path = _work_file_path(work_id)
    with open(path, "w", encoding="utf-8") as f:
        # compact: 대용량 작업 저장·수동 매칭 동기화 시 디스크·네트워크 비용 절감
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))


def _remove_saved_work(work_id: str) -> bool:
    path = _work_file_path(work_id)
    if os.path.isfile(path):
        try:
            os.remove(path)
            return True
        except Exception:
            pass
    return False


# ---------- 설정(필터·동 외곽) 저장/불러오기 (서버) ----------
SAVED_SETTINGS_DIR = os.path.join(_PROJECT_ROOT, "data", "saved_settings")
SAVED_SETTINGS_INDEX = "index.json"


def _ensure_saved_settings_dir() -> str:
    os.makedirs(SAVED_SETTINGS_DIR, exist_ok=True)
    return SAVED_SETTINGS_DIR


def _normalize_settings_context_project(context_project: Optional[str]) -> str:
    return normalize_project_name(context_project) if str(context_project or "").strip() else ""


def _normalize_settings_context_project_id(context_project_id: Optional[str]) -> str:
    return str(context_project_id or "").strip()


def _settings_file_key(
    project_name: str,
    version_id: str,
    context_project: Optional[str] = None,
    context_project_id: Optional[str] = None,
) -> str:
    import hashlib
    context = _normalize_settings_context_project(context_project)
    context_id = _normalize_settings_context_project_id(context_project_id)
    raw = f"{context_id}::{context}::{project_name}::{version_id}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:24]


def _read_saved_settings_index() -> List[Dict[str, Any]]:
    _ensure_saved_settings_dir()
    path = os.path.join(SAVED_SETTINGS_DIR, SAVED_SETTINGS_INDEX)
    if not os.path.isfile(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        items = data if isinstance(data, list) else []
        return [apply_setting_defaults(entry) for entry in items]
    except Exception:
        return []


def _write_saved_settings_index(items: List[Dict[str, Any]]) -> None:
    _ensure_saved_settings_dir()
    path = os.path.join(SAVED_SETTINGS_DIR, SAVED_SETTINGS_INDEX)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)


def _get_saved_setting(
    project_name: str,
    version_id: str,
    context_project: Optional[str] = None,
    context_project_id: Optional[str] = None,
) -> Optional[Dict[str, Any]]:
    key = _settings_file_key(project_name, version_id, context_project, context_project_id)
    path = os.path.join(SAVED_SETTINGS_DIR, f"{key}.json")
    if not os.path.isfile(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            payload = json.load(f)
        return apply_setting_defaults(payload if isinstance(payload, dict) else {})
    except Exception:
        return None


def _put_saved_setting(
    project_name: str,
    version_id: str,
    data: Dict[str, Any],
    context_project: Optional[str] = None,
    context_project_id: Optional[str] = None,
) -> None:
    _ensure_saved_settings_dir()
    context = _normalize_settings_context_project(context_project)
    context_id = _normalize_settings_context_project_id(context_project_id)
    key = _settings_file_key(project_name, version_id, context, context_id)
    path = os.path.join(SAVED_SETTINGS_DIR, f"{key}.json")
    data = apply_setting_defaults(dict(data))
    data["projectName"] = normalize_project_name(project_name)
    data["versionId"] = version_id
    if context:
        data["contextProject"] = context
    else:
        data.pop("contextProject", None)
    if context_id:
        data["contextProjectId"] = context_id
    else:
        data.pop("contextProjectId", None)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _remove_saved_setting(
    project_name: str,
    version_id: str,
    context_project: Optional[str] = None,
    context_project_id: Optional[str] = None,
) -> bool:
    key = _settings_file_key(project_name, version_id, context_project, context_project_id)
    path = os.path.join(SAVED_SETTINGS_DIR, f"{key}.json")
    if os.path.isfile(path):
        try:
            os.remove(path)
            return True
        except Exception:
            pass
    return False


app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=[
        "X-Ortho-Full-W",
        "X-Ortho-Full-H",
        "X-Ortho-Src-X0",
        "X-Ortho-Src-Y0",
        "X-Ortho-Src-W",
        "X-Ortho-Src-H",
        "X-Ortho-Out-W",
        "X-Ortho-Out-H",
        "X-Ortho-Source",
        "X-Ortho-Max-Edge",
    ],
)


@app.middleware("http")
async def add_no_cache_headers(request, call_next):
    response = await call_next(request)
    path = request.url.path
    if path == "/" or path.startswith("/frontend/"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response

app.mount(
    "/frontend",
    StaticFiles(directory=os.path.join(_PROJECT_ROOT, "frontend")),
    name="frontend",
)

app.include_router(pile_dataset_router)

FRONTEND_INDEX_PATH = os.path.join(_PROJECT_ROOT, "frontend", "index.html")

last_raw_circles: Optional[List[Dict]] = None
last_raw_texts: Optional[List[Dict]] = None
last_raw_polylines: Optional[List[Dict]] = None
last_result: Optional[CircleResponse] = None
last_dxf_file_path: Optional[str] = None  # 업로드된 DXF 파일 경로 저장
building_definitions: List[BuildingDefinition] = []
current_filters: Optional[FilterSettings] = None
manual_overrides: Dict[str, str] = {}
history_manual_overrides: Dict[str, str] = {}
last_match_corrections: List[Dict[str, Any]] = []
expected_building_clusters: Optional[int] = None
clustering_max_distance_from_seed: Optional[float] = None
clustering_merge_seed_distance: Optional[float] = None


def set_expected_building_clusters(value: Optional[int]) -> None:
    global expected_building_clusters
    if value is None:
        expected_building_clusters = None
        return
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        expected_building_clusters = None
        return
    expected_building_clusters = parsed if parsed > 0 else None


def merged_manual_overrides() -> Dict[str, str]:
    """히스토리 자동 재적용 + 이 세션 수동 매칭(세션이 우선)."""
    merged = dict(history_manual_overrides)
    merged.update(manual_overrides)
    return merged


def refresh_manual_history_overrides(
    project_context: Optional[str],
    source_type: Optional[str],
    *,
    reuse: bool = True,
    history_reference_source_type: Optional[str] = None,
    history_reference_work_id: Optional[str] = None,
    clear_session_overrides: bool = False,
) -> None:
    global manual_overrides, history_manual_overrides, last_match_corrections
    if clear_session_overrides:
        manual_overrides = {}
    history_manual_overrides = {}
    last_match_corrections = []
    if not reuse:
        return
    if (
        last_raw_circles is None
        or last_raw_texts is None
        or not last_raw_circles
        or not last_raw_texts
    ):
        return
    ref_work = str(history_reference_work_id or "").strip() or None
    if ref_work:
        saved = _get_saved_work(ref_work)
        if not saved:
            return
        history_project = normalize_project_name(saved.get("project"))
        history_source = normalize_source_type(saved.get("sourceType"))
    else:
        if not project_context:
            return
        history_project = normalize_project_name(project_context)
        history_source = (
            normalize_source_type(history_reference_source_type)
            if history_reference_source_type
            else normalize_source_type(source_type)
        )

    history_manual_overrides, last_match_corrections = apply_manual_history(
        history_project,
        history_source,
        last_raw_circles,
        last_raw_texts,
        reference_work_id=ref_work,
    )


def validate_range(name: str, min_value: float, max_value: float) -> Tuple[float, float]:
    if min_value is None or max_value is None:
        raise HTTPException(status_code=400, detail=f"{name} range requires both min and max.")
    if min_value >= max_value:
        raise HTTPException(
            status_code=400,
            detail=f"{name} minimum must be smaller than maximum.",
        )
    return min_value, max_value


def resolve_filter_values(
    min_diameter: Optional[float],
    max_diameter: Optional[float],
    min_area: Optional[float],
    max_area: Optional[float],
    text_height_min: Optional[float],
    text_height_max: Optional[float],
    max_match_distance: Optional[float] = None,
    text_reference_point: Optional[str] = None,
) -> FilterSettings:
    if last_result:
        current = last_result.filter
    else:
        current = FilterSettings(
            min_diameter=DEFAULT_MIN_DIAMETER,
            max_diameter=DEFAULT_MAX_DIAMETER,
            min_area=None,
            max_area=None,
            text_height_min=DEFAULT_TEXT_HEIGHT_MIN,
            text_height_max=DEFAULT_TEXT_HEIGHT_MAX,
            max_match_distance=DEFAULT_MAX_MATCH_DISTANCE,
            text_reference_point="center",
        )

    target = FilterSettings(
        min_diameter=min_diameter if min_diameter is not None else current.min_diameter,
        max_diameter=max_diameter if max_diameter is not None else current.max_diameter,
        min_area=min_area if min_area is not None else current.min_area,
        max_area=max_area if max_area is not None else current.max_area,
        text_height_min=(
            text_height_min if text_height_min is not None else current.text_height_min
        ),
        text_height_max=(
            text_height_max if text_height_max is not None else current.text_height_max
        ),
        max_match_distance=(
            max_match_distance
            if max_match_distance is not None
            else current.max_match_distance
        ),
        text_reference_point=(
            text_reference_point
            if text_reference_point is not None
            else current.text_reference_point
        ),
    )

    validate_range("Diameter", target.min_diameter, target.max_diameter)
    if target.min_area is not None and target.max_area is not None:
        validate_range("Area", target.min_area, target.max_area)
    validate_range("Text height", target.text_height_min, target.text_height_max)
    if target.max_match_distance <= 0:
        raise HTTPException(
            status_code=400, detail="Max match distance must be greater than zero."
        )
    return target


def _circle_area(circle: Dict) -> float:
    """원의 면적 반환. 없으면 radius로 계산."""
    if "area" in circle and circle["area"] is not None:
        return float(circle["area"])
    r = circle.get("radius")
    if r is not None:
        return math.pi * float(r) * float(r)
    d = circle.get("diameter")
    if d is not None:
        return math.pi * (float(d) / 2.0) ** 2
    return 0.0


def filter_circles(
    circles: List[Dict],
    min_diameter: float,
    max_diameter: float,
    min_area: Optional[float] = None,
    max_area: Optional[float] = None,
) -> List[Dict]:
    """직경 범위 또는 면적 범위에 맞으면 통과. 면적 범위가 있으면 직경 통과 또는 면적 통과 시 포함(원형 폴리라인 포함)."""
    result = []
    use_area = min_area is not None and max_area is not None and min_area < max_area
    for circle in circles:
        c = circle.copy()
        if "area" not in c or c.get("area") is None:
            c["area"] = _circle_area(c)
        d = c["diameter"]
        area = c["area"]
        by_diameter = min_diameter < d < max_diameter
        by_area = use_area and (min_area < area < max_area)
        if by_diameter or by_area:
            result.append(c)
    return result


def _is_foundation_pf_text_item(text: Any) -> bool:
    """구버전 저장 데이터(플래그 누락)까지 포함해 P/F 기초 표기 여부를 판별."""
    if isinstance(text, dict):
        if text.get("foundation_pf_only"):
            return True
        if foundation_pf_only_flag(str(text.get("text") or "")):
            text["foundation_pf_only"] = True
            return True
        return False
    if bool(getattr(text, "foundation_pf_only", False)):
        return True
    return foundation_pf_only_flag(str(getattr(text, "text", "") or ""))


def _count_non_foundation_texts(texts: List[Any]) -> int:
    return sum(1 for text in texts or [] if not _is_foundation_pf_text_item(text))


def filter_texts(
    texts: List[Dict],
    min_height: float,
    max_height: float,
) -> List[Dict]:
    filtered: List[Dict] = []
    for text in texts:
        copied = text.copy()
        is_pf_only = _is_foundation_pf_text_item(copied)
        if min_height <= copied["height"] <= max_height or is_pf_only:
            filtered.append(copied)
    return filtered


def build_text_spatial_index(
    texts: List[Dict],
    cell_size: float,
    text_reference_point: str = "center",
) -> Dict[Tuple[int, int], List[Dict]]:
    index: Dict[Tuple[int, int], List[Dict]] = {}
    use_insert = text_reference_point == "insert"
    for text in texts:
        # 기준점에 따라 사용할 좌표 결정
        tx = text["insert_x"] if use_insert else text["text_center_x"]
        ty = text["insert_y"] if use_insert else text["text_center_y"]
        cell = (
            int(math.floor(tx / cell_size)),
            int(math.floor(ty / cell_size)),
        )
        index.setdefault(cell, []).append(text)
    return index


def iter_neighbor_cells(cx: int, cy: int, radius: int) -> List[Tuple[int, int]]:
    if radius == 0:
        return [(cx, cy)]
    cells = []
    for dx in range(-radius, radius + 1):
        for dy in range(-radius, radius + 1):
            cells.append((cx + dx, cy + dy))
    return cells


def gather_text_candidates(
    circle: Dict,
    spatial_index: Dict[Tuple[int, int], List[Dict]],
    texts: List[Dict],
    max_distance: Optional[float],
    text_reference_point: str = "center",
    grid_radius: int = MAX_SEARCH_RADIUS,
) -> tuple[List[Tuple[Dict, float]], bool]:
    cx = circle["center_x"]
    cy = circle["center_y"]
    cell_x = int(math.floor(cx / GRID_CELL_SIZE))
    cell_y = int(math.floor(cy / GRID_CELL_SIZE))
    candidates: List[Tuple[Dict, float]] = []
    seen_ids: set[str] = set()
    
    # 텍스트 기준점에 따라 사용할 좌표 결정
    use_insert = text_reference_point == "insert"

    # 이전: 반지름을 키우다가 후보가 생기면 즉시 중단 → 인접 격자의 올바른 텍스트가
    # 후보에서 빠져 그리디/최소비용 유량이 국소 최적에 갇히는 문제(영통 5동 등).
    for nx, ny in iter_neighbor_cells(cell_x, cell_y, grid_radius):
        for text in spatial_index.get((nx, ny), []):
            text_id = text["id"]
            if text_id in seen_ids:
                continue
            seen_ids.add(text_id)
            tx = text["insert_x"] if use_insert else text["text_center_x"]
            ty = text["insert_y"] if use_insert else text["text_center_y"]
            distance = math.hypot(cx - tx, cy - ty)
            candidates.append((text, distance))

    if not candidates:
        for text in texts:
            text_id = text["id"]
            if text_id in seen_ids:
                continue
            seen_ids.add(text_id)
            # 기준점에 따라 좌표 선택
            tx = text["insert_x"] if use_insert else text["text_center_x"]
            ty = text["insert_y"] if use_insert else text["text_center_y"]
            distance = math.hypot(cx - tx, cy - ty)
            candidates.append((text, distance))

    candidates.sort(key=lambda item: item[1])
    had_any = bool(candidates)
    if max_distance is not None and max_distance > 0:
        candidates = [item for item in candidates if item[1] <= max_distance]
    return candidates, had_any


def _circle_match_priority(circle: Dict[str, Any]) -> int:
    if circle.get("source") == "POLYLINE":
        return 1
    return 0


def _unit_vector_circle_to_text(
    circle: Dict[str, Any],
    text: Dict[str, Any],
    text_reference_point: str,
) -> Optional[Tuple[float, float]]:
    use_insert = text_reference_point == "insert"
    tx = text["insert_x"] if use_insert else text["text_center_x"]
    ty = text["insert_y"] if use_insert else text["text_center_y"]
    wx = float(tx) - float(circle["center_x"])
    wy = float(ty) - float(circle["center_y"])
    d = math.hypot(wx, wy)
    if d < 1e-12:
        return None
    return wx / d, wy / d


def _manual_override_preferred_direction(
    overrides: Dict[str, str],
    circle_lookup: Dict[str, Dict[str, Any]],
    text_lookup: Dict[str, Dict[str, Any]],
    text_reference_point: str,
) -> Optional[Tuple[float, float]]:
    if not overrides:
        return None
    sx, sy = 0.0, 0.0
    n = 0
    for circle_id, text_id in overrides.items():
        circle = circle_lookup.get(circle_id)
        text = text_lookup.get(text_id)
        if not circle or not text:
            continue
        u = _unit_vector_circle_to_text(circle, text, text_reference_point)
        if not u:
            continue
        sx += u[0]
        sy += u[1]
        n += 1
    if n == 0:
        return None
    norm = math.hypot(sx, sy)
    if norm < 1e-12:
        return None
    return sx / norm, sy / norm


def _merge_component_preferred_direction(
    manual_pref: Optional[Tuple[float, float]],
    component_circle_ids: List[str],
    circle_states: Dict[str, Dict[str, Optional[object]]],
    circle_lookup: Dict[str, Dict[str, Any]],
    text_reference_point: str,
) -> Optional[Tuple[float, float]]:
    """컴포넌트별 선호 단위벡터: 수동 매칭 대세 + 각 원의 최근접 후보 방향(거리 역가중)."""
    vx, vy = 0.0, 0.0
    if manual_pref is not None and LABEL_DIRECTION_MANUAL_VOTE_EQUIV > 0:
        mw = float(LABEL_DIRECTION_MANUAL_VOTE_EQUIV)
        vx += manual_pref[0] * mw
        vy += manual_pref[1] * mw
    for cid in component_circle_ids:
        state = circle_states.get(cid)
        if not state:
            continue
        cands = state.get("candidates") or []
        if not cands:
            continue
        text_data, dist = min(cands, key=lambda x: x[1])
        circle = circle_lookup[cid]
        u = _unit_vector_circle_to_text(circle, text_data, text_reference_point)
        if not u:
            continue
        w = 1.0 / (float(dist) + 1e-3)
        vx += u[0] * w
        vy += u[1] * w
    norm = math.hypot(vx, vy)
    if norm < 1e-12:
        return None
    return vx / norm, vy / norm


def _label_direction_penalty(
    circle: Dict[str, Any],
    text: Dict[str, Any],
    text_reference_point: str,
    pref: Optional[Tuple[float, float]],
) -> int:
    if LABEL_DIRECTION_ALIGN_COST_MAX <= 0 or pref is None:
        return 0
    pref_ux, pref_uy = pref[0], pref[1]
    use_insert = text_reference_point == "insert"
    tx = text["insert_x"] if use_insert else text["text_center_x"]
    ty = text["insert_y"] if use_insert else text["text_center_y"]
    wx = float(tx) - float(circle["center_x"])
    wy = float(ty) - float(circle["center_y"])
    d = math.hypot(wx, wy)
    if d < 1e-12:
        return 0
    cos_al = (wx * pref_ux + wy * pref_uy) / d
    cos_al = max(-1.0, min(1.0, cos_al))
    misalign = 0.5 * (1.0 - cos_al)
    return int(LABEL_DIRECTION_ALIGN_COST_MAX * misalign)


def _build_match_edge_cost(
    distance: float,
    circle: Dict[str, Any],
    text_data: Dict[str, Any],
    text_reference_point: str,
    pref: Optional[Tuple[float, float]],
) -> int:
    # 거리 우선, 동일 거리에서는 폴리라인 원형 텍스트를 먼저 유지한다.
    base = int(round(distance * MATCH_COST_SCALE)) * 2 + (
        0 if _circle_match_priority(circle) else 1
    )
    return base + _label_direction_penalty(circle, text_data, text_reference_point, pref)


def _build_matching_components(
    circle_states: Dict[str, Dict[str, Optional[object]]],
) -> List[Tuple[List[str], List[str]]]:
    text_to_circles: Dict[str, List[str]] = defaultdict(list)
    for circle_id, state in circle_states.items():
        for text_data, _distance in state["candidates"]:
            text_to_circles[text_data["id"]].append(circle_id)

    components: List[Tuple[List[str], List[str]]] = []
    visited_circles: set[str] = set()
    visited_texts: set[str] = set()

    for circle_id, state in circle_states.items():
        if circle_id in visited_circles or not state["candidates"]:
            continue
        component_circles: List[str] = []
        component_texts: List[str] = []
        queue: deque[Tuple[str, str]] = deque([("circle", circle_id)])
        visited_circles.add(circle_id)

        while queue:
            kind, node_id = queue.popleft()
            if kind == "circle":
                component_circles.append(node_id)
                circle_state = circle_states[node_id]
                for text_data, _distance in circle_state["candidates"]:
                    text_id = text_data["id"]
                    if text_id in visited_texts:
                        continue
                    visited_texts.add(text_id)
                    queue.append(("text", text_id))
            else:
                component_texts.append(node_id)
                for linked_circle_id in text_to_circles.get(node_id, []):
                    if linked_circle_id in visited_circles:
                        continue
                    visited_circles.add(linked_circle_id)
                    queue.append(("circle", linked_circle_id))

        components.append((component_circles, component_texts))

    return components


def _solve_matching_component(
    component_circle_ids: List[str],
    component_text_ids: List[str],
    circle_states: Dict[str, Dict[str, Optional[object]]],
    circle_lookup: Dict[str, Dict[str, Any]],
    text_reference_point: str = "center",
    manual_pref: Optional[Tuple[float, float]] = None,
) -> Dict[str, Tuple[Dict[str, Any], float]]:
    if not component_circle_ids or not component_text_ids:
        return {}

    class FlowEdge:
        __slots__ = ("to", "rev", "capacity", "cost")

        def __init__(self, to: int, rev: int, capacity: int, cost: int) -> None:
            self.to = to
            self.rev = rev
            self.capacity = capacity
            self.cost = cost

    def add_edge(graph: List[List[FlowEdge]], source: int, target: int, capacity: int, cost: int) -> None:
        forward = FlowEdge(target, len(graph[target]), capacity, cost)
        backward = FlowEdge(source, len(graph[source]), 0, -cost)
        graph[source].append(forward)
        graph[target].append(backward)

    source = 0
    sink = 1
    next_node = 2
    circle_nodes: Dict[str, int] = {}
    text_nodes: Dict[str, int] = {}

    for circle_id in component_circle_ids:
        circle_nodes[circle_id] = next_node
        next_node += 1
    for text_id in component_text_ids:
        text_nodes[text_id] = next_node
        next_node += 1

    graph: List[List[FlowEdge]] = [[] for _ in range(next_node)]
    candidate_lookup: Dict[Tuple[str, str], Tuple[Dict[str, Any], float]] = {}

    pref = _merge_component_preferred_direction(
        manual_pref,
        component_circle_ids,
        circle_states,
        circle_lookup,
        text_reference_point,
    )

    for circle_id, node in circle_nodes.items():
        add_edge(graph, source, node, 1, 0)
        circle = circle_lookup[circle_id]
        for text_data, distance in circle_states[circle_id]["candidates"]:
            text_id = text_data["id"]
            if text_id not in text_nodes:
                continue
            candidate_lookup[(circle_id, text_id)] = (text_data, distance)
            add_edge(
                graph,
                node,
                text_nodes[text_id],
                1,
                _build_match_edge_cost(
                    distance, circle, text_data, text_reference_point, pref
                ),
            )

    for text_id, node in text_nodes.items():
        add_edge(graph, node, sink, 1, 0)

    potentials = [0] * len(graph)

    while True:
        distances = [math.inf] * len(graph)
        previous_node = [-1] * len(graph)
        previous_edge = [-1] * len(graph)
        distances[source] = 0
        heap: List[Tuple[float, int]] = [(0, source)]

        while heap:
            current_distance, node = heapq.heappop(heap)
            if current_distance != distances[node]:
                continue
            for edge_index, edge in enumerate(graph[node]):
                if edge.capacity <= 0:
                    continue
                reduced_cost = edge.cost + potentials[node] - potentials[edge.to]
                next_distance = current_distance + reduced_cost
                if next_distance >= distances[edge.to]:
                    continue
                distances[edge.to] = next_distance
                previous_node[edge.to] = node
                previous_edge[edge.to] = edge_index
                heapq.heappush(heap, (next_distance, edge.to))

        if math.isinf(distances[sink]):
            break

        for node, value in enumerate(distances):
            if not math.isinf(value):
                potentials[node] += int(value)

        node = sink
        while node != source:
            parent = previous_node[node]
            edge = graph[parent][previous_edge[node]]
            edge.capacity -= 1
            graph[node][edge.rev].capacity += 1
            node = parent

    assigned_text_nodes = {node: text_id for text_id, node in text_nodes.items()}
    assignments: Dict[str, Tuple[Dict[str, Any], float]] = {}
    for circle_id, node in circle_nodes.items():
        for edge in graph[node]:
            text_id = assigned_text_nodes.get(edge.to)
            if not text_id or edge.capacity != 0:
                continue
            candidate = candidate_lookup.get((circle_id, text_id))
            if candidate is not None:
                assignments[circle_id] = candidate
            break

    return assignments


def point_in_polygon(x: float, y: float, vertices: List[Tuple[float, float]]) -> bool:
    inside = False
    n = len(vertices)
    if n < 3:
        return False
    j = n - 1
    for i in range(n):
        xi, yi = vertices[i]
        xj, yj = vertices[j]
        intersects = ((yi > y) != (yj > y)) and (
            x < (xj - xi) * (y - yi) / (yj - yi + 1e-12) + xi
        )
        if intersects:
            inside = not inside
        j = i
    return inside


# Fallback outline(polyline) should roughly follow pile spread.
# Extremely large site boundaries can swallow all texts/circles and break mapping.
AUTO_POLYLINE_MAX_AREA_TO_CIRCLE_BBOX_RATIO = 20.0


def _polygon_area(vertices: List[Tuple[float, float]]) -> float:
    if len(vertices) < 3:
        return 0.0
    area = 0.0
    n = len(vertices)
    for i in range(n):
        x1, y1 = vertices[i]
        x2, y2 = vertices[(i + 1) % n]
        area += x1 * y2 - x2 * y1
    return abs(area) * 0.5


def _circles_bbox_area(circles: List[Dict]) -> Optional[float]:
    if not circles:
        return None
    xs = [float(circle.get("center_x", 0.0)) for circle in circles]
    ys = [float(circle.get("center_y", 0.0)) for circle in circles]
    width = max(xs) - min(xs)
    height = max(ys) - min(ys)
    area = width * height
    return area if area > 0 else None


def _is_excessive_fallback_polyline(
    vertices: List[Tuple[float, float]],
    circles_bbox_area: Optional[float],
) -> bool:
    if circles_bbox_area is None or circles_bbox_area <= 0:
        return False
    area = _polygon_area(vertices)
    if area <= 0:
        return False
    return area > circles_bbox_area * AUTO_POLYLINE_MAX_AREA_TO_CIRCLE_BBOX_RATIO


def _polyline_vertices(polyline: Dict[str, Any]) -> List[Tuple[float, float]]:
    points = polyline.get("points")
    if not isinstance(points, list):
        return []
    vertices: List[Tuple[float, float]] = []
    for point in points:
        if not isinstance(point, dict):
            continue
        if "x" not in point or "y" not in point:
            continue
        try:
            vertices.append((float(point["x"]), float(point["y"])))
        except (TypeError, ValueError):
            continue
    return vertices


def _sanitize_reference_polylines(
    polylines: Optional[List[Dict]],
    circles: List[Dict],
) -> List[Dict]:
    rows = list(polylines or [])
    if not rows:
        return []
    circles_bbox_area = _circles_bbox_area(circles)
    sanitized: List[Dict] = []
    for index, polyline in enumerate(rows):
        if not isinstance(polyline, dict):
            continue
        vertices = _polyline_vertices(polyline)
        if (
            len(vertices) >= 3
            and polyline.get("closed") is not False
            and _is_excessive_fallback_polyline(vertices, circles_bbox_area)
        ):
            logger.debug(
                "Dropping oversized reference polyline from response: %s",
                polyline.get("id") or f"poly_{index}",
            )
            continue
        sanitized.append(polyline)
    return sanitized



def classify_entities(
    circles: List[Dict],
    texts: List[Dict],
    buildings: List[BuildingDefinition],
    polylines: Optional[List[Dict]] = None,
) -> None:
    """
    Assign circles/texts to building polygons. Prefer user-defined buildings and
    fall back to clustering polygons when necessary.
    """
    polygons: List[Tuple[int, int, str, List[Tuple[float, float]]]] = []
    circles_bbox_area = _circles_bbox_area(circles)

    # Use building definitions first
    for index, building in enumerate(buildings):
        if len(building.vertices) >= 3:
            kind = str(getattr(building, "kind", "building") or "building").strip().lower()
            # 낮은 숫자가 먼저 검사되어 할당됨. 타워크레인은 지하주차장·일반 동보다 우선(겹침 시 타워 윤곽).
            if kind == "tower_crane":
                priority = 0
            elif kind == "parking":
                priority = 2
            else:
                priority = 1
            polygons.append(
                (
                    priority,
                    index,
                    building.name,
                    [(vertex.x, vertex.y) for vertex in building.vertices],
                )
            )
    polygons.sort(key=lambda item: (item[0], item[1]))

    # Fallback to clustering polygons only when no buildings are defined
    if not polygons and polylines:
        for idx, polyline in enumerate(polylines):
            cluster_type = polyline.get("cluster_type")
            if cluster_type and cluster_type != "building":
                continue
            if polyline.get("closed") is False:
                continue
            points = polyline.get("points")
            if not points or len(points) < 3:
                continue

            vertices = [
                (p.get("x", 0), p.get("y", 0))
                for p in points
                if "x" in p and "y" in p
            ]
            if len(vertices) >= 3:
                if _is_excessive_fallback_polyline(vertices, circles_bbox_area):
                    logger.debug(
                        "Skipping oversized fallback polyline: %s (area=%.2f, circle_bbox_area=%.2f)",
                        polyline.get("id") or f"polyline_{idx + 1}",
                        _polygon_area(vertices),
                        circles_bbox_area,
                    )
                    continue
                polyline_name = (
                    polyline.get("cluster_id")
                    or polyline.get("id")
                    or f"building_{idx + 1}"
                )
                polygons.append((0, idx, polyline_name, vertices))
                logger.debug(
                    "Using cluster polygon: %s (%d vertices)",
                    polyline_name,
                    len(vertices),
                )

    # No polygons -> nothing to assign
    if not polygons:
        for circle in circles:
            circle["building_name"] = None
            circle["building_seq"] = None
        for text in texts:
            text["building_name"] = None
        return

    assigned_circles: Dict[str, List[Dict]] = defaultdict(list)

    for circle in circles:
        circle["building_name"] = None
        circle["building_seq"] = None
        for _, _, name, polygon in polygons:
            if point_in_polygon(circle["center_x"], circle["center_y"], polygon):
                circle["building_name"] = name
                assigned_circles[name].append(circle)
                break

    for name, members in assigned_circles.items():
        members.sort(key=lambda item: (-item["center_y"], item["center_x"]))
        for index, circle in enumerate(members, start=1):
            circle["building_seq"] = index

    for text in texts:
        text["building_name"] = None
        for _, _, name, polygon in polygons:
            if point_in_polygon(text["text_center_x"], text["text_center_y"], polygon):
                text["building_name"] = name
                break

def extract_label_and_number(circle: Dict) -> Tuple[Optional[str], Optional[int]]:
    matched = circle.get("matched_text")
    raw_value: Optional[str] = None
    if isinstance(matched, dict):
        raw_value = matched.get("text")
    elif hasattr(matched, "text"):
        raw_value = getattr(matched, "text")
    if raw_value is None:
        return None, None
    candidate = str(raw_value).strip()
    if not candidate or not NUMERIC_LABEL_PATTERN.fullmatch(candidate):
        return None, None
    try:
        return candidate, int(candidate)
    except ValueError:
        return None, None


def build_pile_points_for_clustering(circles: List[Dict]) -> List[PilePoint]:
    points: List[PilePoint] = []
    for circle in circles:
        label, number = extract_label_and_number(circle)
        points.append(
            PilePoint(
                id=circle["id"],
                x=float(circle["center_x"]),
                y=float(circle["center_y"]),
                label=label,
                number=number,
                layer=circle.get("layer"),
            )
        )
    return points


def build_cluster_payload(cluster: PileCluster) -> Dict:
    return {
        "id": cluster.id,
        "type": cluster.type,
        "points": [
            {
                "id": point.id,
                "x": point.x,
                "y": point.y,
                "label": point.label,
                "number": point.number,
            }
            for point in cluster.points
        ],
        "centroid_x": cluster.centroid[0],
        "centroid_y": cluster.centroid[1],
        "hull": [{"x": x, "y": y} for x, y in cluster.hull],
        "metrics": {
            "count_points": cluster.metrics.count_points,
            "min_num": cluster.metrics.min_num,
            "max_num": cluster.metrics.max_num,
            "range_size": cluster.metrics.range_size,
            "missing_count": cluster.metrics.missing_count,
            "missing_ratio": cluster.metrics.missing_ratio,
            "density": cluster.metrics.density,
            "bbox_area": cluster.metrics.bbox_area,
            "width": cluster.metrics.width,
            "height": cluster.metrics.height,
        },
        "score": cluster.score,
    }


def clusters_to_polylines(clusters: List[PileCluster]) -> List[Dict]:
    polylines: List[Dict] = []
    for cluster in clusters:
        if cluster.type != "building":
            continue
        if not cluster.hull or len(cluster.hull) < 3:
            continue
        polylines.append(
            {
                "id": f"{cluster.id}_poly",
                "closed": True,
                "points": [{"x": x, "y": y} for x, y in cluster.hull],
                "layer": cluster.type,
                "cluster_id": cluster.id,
                "cluster_type": cluster.type,
            }
        )
    return polylines


def compute_building_summary(
    circles: List[Dict],
    texts: List[Dict],
    buildings: List[BuildingDefinition],
) -> List[BuildingSummary]:
    summary: Dict[str, Dict[str, int]] = defaultdict(
        lambda: {"circles": 0, "texts": 0, "errors": 0}
    )
    for circle in circles:
        name = circle.get("building_name") or "미할당"
        summary[name]["circles"] += 1
        if circle.get("has_error"):
            summary[name]["errors"] += 1
    for text in texts:
        if _is_foundation_pf_text_item(text):
            continue
        name = text.get("building_name") or "미할당"
        summary[name]["texts"] += 1
    for building in buildings:
        label = building.name or "Unnamed"
        summary.setdefault(label, {"circles": 0, "texts": 0, "errors": 0})
    return [
        BuildingSummary(
            name=name,
            circle_count=counts["circles"],
            text_count=counts["texts"],
            error_count=counts["errors"],
        )
        for name, counts in summary.items()
    ]


def _circle_diameter_value(circle: Dict[str, Any]) -> Optional[float]:
    d = circle.get("diameter")
    if d is not None:
        fd = float(d)
        if fd > 0:
            return fd
    r = circle.get("radius")
    if r is not None:
        fr = float(r)
        if fr > 0:
            return fr * 2.0
    return None


def _reference_diameter_for_outlier(
    circle: Dict[str, Any],
    component_circles: List[Dict[str, Any]],
) -> float:
    phi = _circle_diameter_value(circle)
    if phi is not None:
        return max(phi, 1e-9)
    vals: List[float] = []
    for co in component_circles:
        v = _circle_diameter_value(co)
        if v is not None:
            vals.append(v)
    if vals:
        return max(statistics.median(vals), 1e-9)
    return DEFAULT_MIN_DIAMETER


def _euclidean_circle_to_text_distance(
    circle: Dict[str, Any],
    text: Dict[str, Any],
    text_reference_point: str,
) -> float:
    use_insert = text_reference_point == "insert"
    tx = text["insert_x"] if use_insert else text["text_center_x"]
    ty = text["insert_y"] if use_insert else text["text_center_y"]
    return float(
        math.hypot(
            float(tx) - float(circle["center_x"]),
            float(ty) - float(circle["center_y"]),
        )
    )


def _prune_pairwise_swap_manual_overrides(
    circles: List[Dict[str, Any]],
    text_lookup: Dict[str, Dict[str, Any]],
    manual_map: Dict[str, str],
    text_reference_point: str,
    *,
    max_center_separation: float,
    min_improvement: float,
) -> Dict[str, str]:
    """
    둘 다 수동 잠금된 인접 원 쌍에서, 할당된 텍스트만 맞바꾸면 거리 합이 min_improvement 이상 줄면
    두 잠금을 제거(자동 매칭에 맡김). 잘못 저장된 인접 오매칭 복구용.
    """
    if len(manual_map) < 2:
        return manual_map
    by_id = {c["id"]: c for c in circles}
    mo = dict(manual_map)
    to_remove: set[str] = set()
    ids = [cid for cid in mo if cid in by_id and mo[cid] in text_lookup]
    for i, c1 in enumerate(ids):
        p1 = by_id[c1]
        for c2 in ids[i + 1 :]:
            p2 = by_id[c2]
            dcc = math.hypot(
                float(p1["center_x"]) - float(p2["center_x"]),
                float(p1["center_y"]) - float(p2["center_y"]),
            )
            if dcc > max_center_separation:
                continue
            t1_id, t2_id = mo[c1], mo[c2]
            t1, t2 = text_lookup[t1_id], text_lookup[t2_id]
            cur = _euclidean_circle_to_text_distance(
                p1, t1, text_reference_point
            ) + _euclidean_circle_to_text_distance(p2, t2, text_reference_point)
            swp = _euclidean_circle_to_text_distance(
                p1, t2, text_reference_point
            ) + _euclidean_circle_to_text_distance(p2, t1, text_reference_point)
            if swp + min_improvement < cur:
                to_remove.add(c1)
                to_remove.add(c2)
                logger.info(
                    "수동 잠금 제거(인접 교환 개선): %s,%s 거리합 %.4f→%.4f",
                    c1,
                    c2,
                    cur,
                    swp,
                )
    for cid in to_remove:
        mo.pop(cid, None)
    return mo


def _is_suspicious_component_match_distance(
    d: float, peer_distances: List[float], reference_diameter: float
) -> bool:
    """
    같은 매칭 컴포넌트에서 이 쌍만 유독 멀 때 재매칭.
    "멀다"는 max_match_distance가 아니라 해당 원의 도면 직경(파일 값) 기준.
    """
    if len(peer_distances) < 2:
        return False
    lo = min(peer_distances)
    hi = max(peer_distances)
    if d < hi - 1e-9:
        return False
    phi = max(float(reference_diameter), 1e-9)
    if d < phi * MATCH_OUTLIER_TO_DIAMETER_RATIO:
        return False
    if lo < 1e-9:
        return True
    return (d / lo) >= MATCH_OUTLIER_PEER_DISTANCE_RATIO


def _suspicious_match_circle_ids(
    components: List[Tuple[List[str], List[str]]],
    assignments: Dict[str, Tuple[Dict[str, Any], float]],
    circle_lookup: Dict[str, Dict[str, Any]],
) -> set[str]:
    bad: set[str] = set()
    for component_circle_ids, _ in components:
        matched = [c for c in component_circle_ids if c in assignments]
        if len(matched) < 2:
            continue
        circle_objs = [circle_lookup[c] for c in matched]
        dists = [assignments[c][1] for c in matched]
        for c in matched:
            d = assignments[c][1]
            ref_phi = _reference_diameter_for_outlier(circle_lookup[c], circle_objs)
            if _is_suspicious_component_match_distance(d, dists, ref_phi):
                bad.add(c)
    return bad


def _expand_circles_touching_suspicious(
    components: List[Tuple[List[str], List[str]]], suspicious: set[str]
) -> set[str]:
    out: set[str] = set()
    for component_circle_ids, _ in components:
        if any(c in suspicious for c in component_circle_ids):
            out.update(component_circle_ids)
    return out


def match_texts_to_circles(
    circles: List[Dict],
    texts: List[Dict],
    max_match_distance: float,
    overrides: Optional[Dict[str, str]] = None,
    text_reference_point: str = "center",
) -> tuple[int, Dict[str, List[str]]]:
    if not circles:
        return 0, {}
    texts = list(texts or [])
    pile_match_texts = [t for t in texts if not t.get("foundation_pf_only")]

    overrides = overrides or {}
    # 클라이언트/이전 단계에서 남은 matched_text·manual_match가 자동 매칭을 오염시키지 않도록 초기화
    for circle in circles:
        circle.pop("matched_text", None)
        circle["matched_text_id"] = None
        circle["matched_text_distance"] = None
        circle["manual_match"] = False
        circle.pop("match_distance_exceeded", None)

    text_lookup = {text["id"]: text for text in texts}
    circle_lookup = {circle["id"]: circle for circle in circles}
    manual_label_pref = _manual_override_preferred_direction(
        overrides, circle_lookup, text_lookup, text_reference_point
    )
    locked_circles: set[str] = set()
    locked_texts: set[str] = set()
    final_links: Dict[str, List[str]] = defaultdict(list)
    matched_count = 0
    
    # 텍스트 기준점에 따라 사용할 좌표 결정
    use_insert = text_reference_point == "insert"

    def make_payload(circle_obj: Dict, text_obj: Dict) -> Dict:
        # 기준점에 따라 거리 계산
        tx = text_obj["insert_x"] if use_insert else text_obj["text_center_x"]
        ty = text_obj["insert_y"] if use_insert else text_obj["text_center_y"]
        distance = math.hypot(
            circle_obj["center_x"] - tx,
            circle_obj["center_y"] - ty,
        )
        return {
            "id": text_obj["id"],
            "text": text_obj["text"],
            "insert_x": text_obj["insert_x"],
            "insert_y": text_obj["insert_y"],
            "insert_z": text_obj["insert_z"],
            "center_x": text_obj["text_center_x"],
            "center_y": text_obj["text_center_y"],
            "text_center_x": text_obj["text_center_x"],
            "text_center_y": text_obj["text_center_y"],
            "height": text_obj["height"],
            "distance": distance,
        }

    for circle_id, text_id in overrides.items():
        circle = circle_lookup.get(circle_id)
        text = text_lookup.get(text_id)
        if not circle or not text:
            continue
        payload = make_payload(circle, text)
        circle["matched_text"] = payload
        circle["matched_text_id"] = text_id
        circle["matched_text_distance"] = payload["distance"]
        circle["manual_match"] = True
        circle.pop("match_distance_exceeded", None)
        final_links[text_id].append(circle_id)
        locked_circles.add(circle_id)
        locked_texts.add(text_id)
        matched_count += 1

    available_texts = [text for text in pile_match_texts if text["id"] not in locked_texts]
    if not available_texts:
        return matched_count, dict(final_links)

    spatial_index = build_text_spatial_index(available_texts, GRID_CELL_SIZE, text_reference_point)
    circle_states: Dict[str, Dict[str, Optional[object]]] = {}

    for circle in circles:
        if circle["id"] in locked_circles:
            continue
        candidates, had_any = gather_text_candidates(
            circle,
            spatial_index,
            available_texts,
            max_match_distance,
            text_reference_point,
            grid_radius=MAX_SEARCH_RADIUS,
        )
        circle_states[circle["id"]] = {
            "candidates": candidates,
            "had_any": had_any,
        }

    assignments: Dict[str, Tuple[Dict[str, Any], float]] = {}
    for refine_pass in range(MATCH_OUTLIER_REFINE_MAX_PASSES + 1):
        components = _build_matching_components(circle_states)
        assignments.clear()
        for component_circle_ids, component_text_ids in components:
            assignments.update(
                _solve_matching_component(
                    component_circle_ids,
                    component_text_ids,
                    circle_states,
                    circle_lookup,
                    text_reference_point,
                    manual_label_pref,
                )
            )
        suspicious = _suspicious_match_circle_ids(components, assignments, circle_lookup)
        if not suspicious or refine_pass >= MATCH_OUTLIER_REFINE_MAX_PASSES:
            break
        expand_ids = _expand_circles_touching_suspicious(components, suspicious)
        for cid in expand_ids:
            if cid in locked_circles:
                continue
            circle = circle_lookup[cid]
            cands, had_any = gather_text_candidates(
                circle,
                spatial_index,
                available_texts,
                max_match_distance,
                text_reference_point,
                grid_radius=EXPANDED_MATCH_GRID_RADIUS,
            )
            circle_states[cid] = {"candidates": cands, "had_any": had_any}

    for circle_id, (text_data, distance) in assignments.items():
        circle = circle_lookup[circle_id]
        payload = make_payload(circle, text_data)
        circle["matched_text"] = payload
        circle["matched_text_id"] = payload["id"]
        circle["matched_text_distance"] = distance
        circle.pop("match_distance_exceeded", None)
        final_links[payload["id"]].append(circle_id)
        matched_count += 1

    for circle_id, state in circle_states.items():
        if circle_lookup.get(circle_id, {}).get("matched_text_id"):
            continue
        circle = circle_lookup.get(circle_id)
        if not circle:
            continue
        if state["had_any"]:
            circle["match_distance_exceeded"] = True

    return matched_count, dict(final_links)


def build_match_errors(
    circles: List[Dict],
    texts: List[Dict],
    text_links: Dict[str, List[str]],
) -> List[MatchError]:
    circle_errors: Dict[str, List[str]] = defaultdict(list)
    errors: List[MatchError] = []

    for text in texts:
        if _is_foundation_pf_text_item(text):
            text["matched_circle_ids"] = []
            text["has_error"] = False
            continue
        circle_ids = text_links.get(text["id"], [])
        text["matched_circle_ids"] = circle_ids
        if len(circle_ids) > 1:
            text["has_error"] = True
            errors.append(
                MatchError(
                    error_type="TEXT_MULTI_MATCH",
                    text_id=text["id"],
                    text_value=text["text"],
                    circle_ids=circle_ids,
                    message=f"TEXT {text['text']} matches {len(circle_ids)} circles.",
                )
            )
            for cid in circle_ids:
                circle_errors[cid].append("TEXT_MULTI_MATCH")
        elif not circle_ids:
            text["has_error"] = True
            errors.append(
                MatchError(
                    error_type="TEXT_NO_MATCH",
                    text_id=text["id"],
                    text_value=text["text"],
                    circle_ids=[],
                    message=f"TEXT {text['text']} has no matching circle.",
                )
            )
        else:
            text["has_error"] = False

    for circle in circles:
        codes = list(circle_errors.get(circle["id"], []))
        if circle.get("match_distance_exceeded"):
            codes.append("MATCH_DISTANCE_EXCEEDED")
            errors.append(
                MatchError(
                    error_type="MATCH_DISTANCE_EXCEEDED",
                    text_id=None,
                    text_value=None,
                    circle_ids=[circle["id"]],
                    message="Matched TEXT exceeds the allowed distance threshold.",
                )
            )
        if not circle.get("matched_text_id"):
            codes.append("CIRCLE_NO_MATCH")
            errors.append(
                MatchError(
                    error_type="CIRCLE_NO_MATCH",
                    text_id=None,
                    text_value=None,
                    circle_ids=[circle["id"]],
                    message=f"Circle {circle['id']} has no matching TEXT.",
                )
            )
        circle["has_error"] = bool(codes)
        circle["error_codes"] = codes

    return errors


def _circles_overlap(a: Dict, b: Dict) -> bool:
    """두 원이 직경(겹침) 기준으로 겹치는지: 중심 거리 <= 반지름 합"""
    ax, ay = float(a.get("center_x", 0)), float(a.get("center_y", 0))
    bx, by = float(b.get("center_x", 0)), float(b.get("center_y", 0))
    ra = float(a.get("radius") or (a.get("diameter") or 0) / 2)
    rb = float(b.get("radius") or (b.get("diameter") or 0) / 2)
    dist = math.sqrt((bx - ax) ** 2 + (by - ay) ** 2)
    return dist <= ra + rb


def build_duplicate_groups(circles: List[Dict]) -> List[DuplicateGroup]:
    """원의 직경 안에 겹쳐 있는 것끼리 중복 그룹으로 묶음 (Union-Find)."""
    n = len(circles)
    parent = list(range(n))

    def find(i: int) -> int:
        if parent[i] != i:
            parent[i] = find(parent[i])
        return parent[i]

    def union(i: int, j: int) -> None:
        pi, pj = find(i), find(j)
        if pi != pj:
            parent[pi] = pj

    for i in range(n):
        for j in range(i + 1, n):
            if _circles_overlap(circles[i], circles[j]):
                union(i, j)

    comp: Dict[int, List[Dict]] = {}
    for i in range(n):
        root = find(i)
        comp.setdefault(root, []).append(circles[i])

    duplicate_groups: List[DuplicateGroup] = []
    for items in comp.values():
        if len(items) < 2:
            continue
        first = items[0]
        x, y = float(first.get("center_x", 0)), float(first.get("center_y", 0))
        duplicate_groups.append(
            DuplicateGroup(
                coord_key={"x": x, "y": y},
                count=len(items),
                circle_ids=[item["id"] for item in items],
                details=[
                    DuplicateDetail(
                        id=item["id"],
                        layer=item["layer"],
                        block_name=item.get("block_name"),
                        matched_text=MatchedText(**item["matched_text"])
                        if item.get("matched_text")
                        else None,
                        has_error=item.get("has_error", False),
                    )
                    for item in items
                ],
            )
        )
    return duplicate_groups


def _matched_number_from_display_text(text_val: Optional[str]) -> Optional[int]:
    """매칭된 TEXT 표시 문자열에서 말뚝 번호 정수 추출 (프론트 getMatchedTextNumber 와 동일: 비숫자 제거)."""
    if text_val is None:
        return None
    digits = re.sub(r"\D", "", str(text_val).strip())
    if not digits:
        return None
    try:
        n = int(digits)
    except ValueError:
        return None
    return n if n >= 1 else None


def extend_same_building_number_duplicates(
    circles: List[Dict],
    errors: List[MatchError],
    duplicate_groups: List[DuplicateGroup],
) -> None:
    """
    같은 동(또는 미할당) 안에서 동일 번호가 2개 이상 매칭된 좌표를
    매칭 에러·중복 좌표 목록에 추가하고 해당 circle에 error_codes를 붙인다.
    classify_entities 이후에 호출해야 building_name이 채워져 있다.
    """
    groups: Dict[Tuple[str, int], List[Dict]] = defaultdict(list)
    for c in circles:
        if not c.get("matched_text_id"):
            continue
        mt = c.get("matched_text")
        raw_text = None
        if isinstance(mt, dict):
            raw_text = mt.get("text")
        num = _matched_number_from_display_text(raw_text)
        if num is None:
            continue
        bname = (c.get("building_name") or "").strip() or "미할당"
        groups[(bname, num)].append(c)

    for (bname, num), items in groups.items():
        if len(items) < 2:
            continue
        circle_ids = [str(item["id"]) for item in items]
        for item in items:
            codes = list(item.get("error_codes") or [])
            if "SAME_BUILDING_NUMBER_DUPLICATE" not in codes:
                codes.append("SAME_BUILDING_NUMBER_DUPLICATE")
            item["error_codes"] = codes
            item["has_error"] = True

        errors.append(
            MatchError(
                error_type="SAME_BUILDING_NUMBER_DUPLICATE",
                text_id=None,
                text_value=str(num),
                circle_ids=circle_ids,
                message=f'"{bname}"(동)에서 번호 {num}이(가) {len(items)}개 좌표에 중복되었습니다.',
            )
        )

        xs = [float(item.get("center_x", 0)) for item in items]
        ys = [float(item.get("center_y", 0)) for item in items]
        cx = sum(xs) / len(xs)
        cy = sum(ys) / len(ys)
        duplicate_groups.append(
            DuplicateGroup(
                coord_key={"x": cx, "y": cy},
                count=len(items),
                circle_ids=circle_ids,
                details=[
                    DuplicateDetail(
                        id=item["id"],
                        layer=item.get("layer", "0"),
                        block_name=item.get("block_name"),
                        matched_text=MatchedText(**item["matched_text"])
                        if isinstance(item.get("matched_text"), dict)
                        else None,
                        has_error=bool(item.get("has_error", False)),
                    )
                    for item in items
                ],
            )
        )


def to_circle_model(circle: Dict) -> CircleRecord:
    matched = None
    if circle.get("matched_text") and isinstance(circle.get("matched_text"), dict):
        try:
            mt = circle["matched_text"]
            matched = MatchedText(
                id=mt.get("id", ""),
                text=mt.get("text", ""),
                insert_x=float(mt.get("insert_x", mt.get("text_center_x", 0))),
                insert_y=float(mt.get("insert_y", mt.get("text_center_y", 0))),
                insert_z=float(mt.get("insert_z", 0)),
                center_x=float(mt.get("center_x", mt.get("text_center_x", 0))),
                center_y=float(mt.get("center_y", mt.get("text_center_y", 0))),
                text_center_x=float(mt.get("text_center_x", mt.get("center_x", 0))),
                text_center_y=float(mt.get("text_center_y", mt.get("center_y", 0))),
                height=float(mt.get("height", 0)),
                distance=float(mt.get("distance", 0)),
            )
        except (TypeError, ValueError, KeyError):
            matched = None
    area = circle.get("area")
    if area is None:
        area = _circle_area(circle)
    return CircleRecord(
        id=str(circle.get("id", "")),
        center_x=float(circle.get("center_x", 0)),
        center_y=float(circle.get("center_y", 0)),
        center_z=float(circle.get("center_z", 0)),
        radius=float(circle.get("radius", 0)),
        diameter=float(circle.get("diameter", circle.get("radius", 0) * 2)),
        area=float(area),
        layer=str(circle.get("layer", "0")),
        block_name=circle.get("block_name"),
        transformed=circle.get("transformed", True),
        matched_text=matched,
        matched_text_id=circle.get("matched_text_id"),
        matched_text_distance=circle.get("matched_text_distance"),
        has_error=circle.get("has_error", False),
        error_codes=circle.get("error_codes", []),
        building_name=circle.get("building_name"),
        building_seq=circle.get("building_seq"),
        manual_match=circle.get("manual_match"),
    )


def to_text_model(text: Dict) -> TextRecord:
    cx = text.get("text_center_x") or text.get("center_x", 0)
    cy = text.get("text_center_y") or text.get("center_y", 0)
    t = text.get("type") or "TEXT"
    if t not in ("TEXT", "MTEXT"):
        t = "TEXT"
    return TextRecord(
        id=str(text.get("id", "")),
        type=t,
        text=text.get("text", ""),
        insert_x=float(text.get("insert_x", cx)),
        insert_y=float(text.get("insert_y", cy)),
        insert_z=float(text.get("insert_z", 0)),
        center_x=float(cx),
        center_y=float(cy),
        text_center_x=float(cx),
        text_center_y=float(cy),
        height=float(text.get("height", 0)),
        rotation_deg=float(text.get("rotation_deg", 0) or 0),
        layer=str(text.get("layer", "0")),
        block_name=text.get("block_name"),
        matched_circle_ids=text.get("matched_circle_ids", []) or [],
        has_error=text.get("has_error", False),
        building_name=text.get("building_name"),
        foundation_pf_only=bool(text.get("foundation_pf_only")),
    )


def to_match_correction_model(item: Dict[str, Any]) -> MatchCorrection:
    return MatchCorrection(
        match_source="manual_history",
        circle_id=str(item.get("circle_id", "")),
        circle_center_x=float(item.get("circle_center_x", 0)),
        circle_center_y=float(item.get("circle_center_y", 0)),
        matched_text_id=str(item.get("matched_text_id", "")),
        matched_text_value=str(item.get("matched_text_value", "")),
        building_name=item.get("building_name"),
        history_work_id=item.get("history_work_id"),
        history_work_title=item.get("history_work_title"),
        history_project=item.get("history_project"),
        history_source_type=item.get("history_source_type"),
    )


def build_response(
    raw_circles: List[Dict],
    raw_texts: List[Dict],
    raw_polylines: Optional[List[Dict]],
    filters: FilterSettings,
) -> CircleResponse:
    correction_models = [to_match_correction_model(item) for item in last_match_corrections]
    filtered_circles = filter_circles(
        raw_circles,
        filters.min_diameter,
        filters.max_diameter,
        filters.min_area,
        filters.max_area,
    )
    filtered_texts = filter_texts(
        raw_texts, filters.text_height_min, filters.text_height_max
    )
    # 수동 매칭된 circle/text는 필터에 걸려도 결과에 포함시켜 매칭이 반영되도록 함
    mo = merged_manual_overrides()
    filtered_circle_ids = {c["id"] for c in filtered_circles}
    filtered_text_ids = {t["id"] for t in filtered_texts}
    for c in raw_circles:
        if c["id"] in mo and c["id"] not in filtered_circle_ids:
            filtered_circles.append(c)
            filtered_circle_ids.add(c["id"])
    for t in raw_texts:
        if t["id"] in mo.values() and t["id"] not in filtered_text_ids:
            filtered_texts.append(t)
            filtered_text_ids.add(t["id"])

    text_lookup_for_prune = {str(t["id"]): t for t in filtered_texts}
    mo = _prune_pairwise_swap_manual_overrides(
        filtered_circles,
        text_lookup_for_prune,
        mo,
        filters.text_reference_point,
        max_center_separation=max(
            2.5,
            float(filters.max_match_distance) * MANUAL_PRUNE_MAX_CENTER_FACTOR,
        ),
        min_improvement=MANUAL_PRUNE_MIN_DISTANCE_IMPROVEMENT,
    )
    matched_pairs, text_links = match_texts_to_circles(
        filtered_circles,
        filtered_texts,
        filters.max_match_distance,
        mo,
        filters.text_reference_point,
    )
    errors = build_match_errors(filtered_circles, filtered_texts, text_links)
    duplicate_groups = build_duplicate_groups(filtered_circles)
    cluster_payload: List[Dict] = []
    cluster_polylines: List[Dict] = []
    if filtered_circles:
        pile_points = build_pile_points_for_clustering(filtered_circles)
        try:
            # 클러스터링 설정값 전달
            cluster_config = DEFAULT_CLUSTER_CONFIG
            cluster_results = cluster_piles(
                pile_points,
                cluster_config,
                expected_building_clusters,
                max_distance_from_seed=clustering_max_distance_from_seed,
                merge_seed_distance=clustering_merge_seed_distance,
            )
        except Exception:
            logger.exception("Failed to cluster pile coordinates.")
            cluster_results = []
        cluster_payload = [build_cluster_payload(cluster) for cluster in cluster_results]
        cluster_polylines = clusters_to_polylines(cluster_results)
    
    # 건물 분류: 사용자 정의 > 클러스터 폴리곤
    classify_entities(
        filtered_circles,
        filtered_texts,
        building_definitions,
        cluster_polylines if cluster_polylines else None,
    )
    extend_same_building_number_duplicates(filtered_circles, errors, duplicate_groups)

    # 로그: building_name 설정 확인
    assigned_count = sum(1 for c in filtered_circles if c.get("building_name"))
    logger.info(
        "동 분류 완료: 총 %d개 중 %d개가 영역 할당 (buildings: %d, cluster_polylines: %d)",
        len(filtered_circles),
        assigned_count,
        len(building_definitions),
        len(cluster_polylines),
    )
    building_summary = compute_building_summary(
        filtered_circles, filtered_texts, building_definitions
    )
    summary = SummaryStats(
        total_circles=len(filtered_circles),
        total_texts=_count_non_foundation_texts(filtered_texts),
        matched_pairs=matched_pairs,
        duplicate_groups=len(duplicate_groups),
    )
    safe_raw_polylines = _sanitize_reference_polylines(raw_polylines, filtered_circles)
    circles = [to_circle_model(circle) for circle in filtered_circles]
    texts = [to_text_model(text) for text in filtered_texts]
    return CircleResponse(
        summary=summary,
        circles=circles,
        texts=texts,
        duplicates=duplicate_groups,
        polylines=_ensure_polyline_ids(safe_raw_polylines),
        buildings=building_definitions,
        pile_clusters=cluster_payload,
        cluster_polylines=_ensure_polyline_ids(cluster_polylines),
        errors=errors,
        filter=filters,
        building_summary=building_summary,
        match_corrections=correction_models,
    )


def update_cached_result(filters: FilterSettings) -> CircleResponse:
    global current_filters
    current_filters = filters
    # None 체크만 수행 (빈 리스트는 허용)
    if last_raw_circles is None or last_raw_texts is None:
        # 빈 결과 반환 (404 에러 방지)
        return CircleResponse(
            summary=SummaryStats(total_circles=0, total_texts=0, matched_pairs=0, duplicate_groups=0),
            circles=[],
            texts=[],
            duplicates=[],
            errors=[],
            filter=filters,
            match_corrections=[],
        )
    result = build_response(last_raw_circles, last_raw_texts, last_raw_polylines, filters)
    global last_result
    last_result = result
    return result


def _empty_circle_response(filters: FilterSettings) -> CircleResponse:
    """빈 필터 결과용 공통 응답."""
    return CircleResponse(
        summary=SummaryStats(total_circles=0, total_texts=0, matched_pairs=0, duplicate_groups=0),
        circles=[],
        texts=[],
        duplicates=[],
        errors=[],
        filter=filters,
        match_corrections=[],
    )


def build_response_light(filters: FilterSettings) -> CircleResponse:
    """
    수동 매칭(추가/해제) 시 사용. 클러스터링·건물 분류는 재사용하고
    매칭·에러·중복만 재계산하여 응답 속도를 높인다.
    """
    global current_filters, last_result
    current_filters = filters
    if last_raw_circles is None or last_raw_texts is None:
        return _empty_circle_response(filters)
    if last_result is None:
        return update_cached_result(filters)

    filtered_circles = filter_circles(
        last_raw_circles,
        filters.min_diameter,
        filters.max_diameter,
        filters.min_area,
        filters.max_area,
    )
    filtered_texts = filter_texts(
        last_raw_texts, filters.text_height_min, filters.text_height_max
    )
    mo = merged_manual_overrides()
    filtered_circle_ids = {c["id"] for c in filtered_circles}
    filtered_text_ids = {t["id"] for t in filtered_texts}
    for c in last_raw_circles:
        if c["id"] in mo and c["id"] not in filtered_circle_ids:
            filtered_circles.append(c)
            filtered_circle_ids.add(c["id"])
    for t in last_raw_texts:
        if t["id"] in mo.values() and t["id"] not in filtered_text_ids:
            filtered_texts.append(t)
            filtered_text_ids.add(t["id"])

    text_lookup_light = {str(t["id"]): t for t in filtered_texts}
    mo = _prune_pairwise_swap_manual_overrides(
        filtered_circles,
        text_lookup_light,
        mo,
        filters.text_reference_point,
        max_center_separation=max(
            2.5,
            float(filters.max_match_distance) * MANUAL_PRUNE_MAX_CENTER_FACTOR,
        ),
        min_improvement=MANUAL_PRUNE_MIN_DISTANCE_IMPROVEMENT,
    )
    matched_pairs, text_links = match_texts_to_circles(
        filtered_circles,
        filtered_texts,
        filters.max_match_distance,
        mo,
        filters.text_reference_point,
    )
    errors = build_match_errors(filtered_circles, filtered_texts, text_links)
    duplicate_groups = build_duplicate_groups(filtered_circles)

    # 이전 결과에서 건물명만 복사 (클러스터 재계산 생략)
    old_circle_by_id = {c.id: c for c in last_result.circles}
    old_text_by_id = {t.id: t for t in last_result.texts}
    for c in filtered_circles:
        old = old_circle_by_id.get(c["id"])
        if old and old.building_name is not None and not is_placeholder_building_name(old.building_name):
            c["building_name"] = old.building_name
        if old and old.building_seq is not None:
            c["building_seq"] = old.building_seq
    for t in filtered_texts:
        old = old_text_by_id.get(t["id"])
        if old and old.building_name is not None and not is_placeholder_building_name(old.building_name):
            t["building_name"] = old.building_name

    extend_same_building_number_duplicates(filtered_circles, errors, duplicate_groups)

    building_summary = compute_building_summary(
        filtered_circles, filtered_texts, building_definitions
    )
    summary = SummaryStats(
        total_circles=len(filtered_circles),
        total_texts=_count_non_foundation_texts(filtered_texts),
        matched_pairs=matched_pairs,
        duplicate_groups=len(duplicate_groups),
    )
    circles = [to_circle_model(c) for c in filtered_circles]
    texts = [to_text_model(t) for t in filtered_texts]
    result = CircleResponse(
        summary=summary,
        circles=circles,
        texts=texts,
        duplicates=duplicate_groups,
        polylines=_ensure_polyline_ids(last_raw_polylines or []),
        buildings=building_definitions,
        pile_clusters=last_result.pile_clusters,
        cluster_polylines=last_result.cluster_polylines,
        errors=errors,
        filter=filters,
        building_summary=building_summary,
        match_corrections=[to_match_correction_model(item) for item in last_match_corrections],
    )
    last_result = result
    return result


def apply_building_definitions(definitions: List[BuildingDefinition]) -> CircleResponse:
    global building_definitions
    building_definitions = definitions
    filters = current_filters or FilterSettings(
        min_diameter=DEFAULT_MIN_DIAMETER,
        max_diameter=DEFAULT_MAX_DIAMETER,
        text_height_min=DEFAULT_TEXT_HEIGHT_MIN,
        text_height_max=DEFAULT_TEXT_HEIGHT_MAX,
        max_match_distance=DEFAULT_MAX_MATCH_DISTANCE,
    )
    return update_cached_result(filters)


def prepare_export_rows(
    circles: List[CircleRecord],
    texts: List[TextRecord],
    errors: List[MatchError],
    buildings: List[BuildingDefinition],
) -> tuple[List[Dict], List[Dict], List[Dict], List[str]]:
    grouped: Dict[Tuple[float, float], List[CircleRecord]] = {}
    for circle in circles:
        key = (
            round(circle.center_x, COORD_PRECISION),
            round(circle.center_y, COORD_PRECISION),
        )
        grouped.setdefault(key, []).append(circle)

    rows: List[Dict] = []
    for items in grouped.values():
        clean_items = [item for item in items if not item.has_error]
        if not clean_items:
            continue
        sorted_items = sorted(
            clean_items,
            key=lambda c: (c.matched_text is None, c.id),
        )
        representative = sorted_items[0]
        rows.append(
            {
                PILE_NO_HEADER: representative.matched_text.text
                if representative.matched_text
                else "",
                PILE_X_HEADER: representative.center_x,
                PILE_Y_HEADER: representative.center_y,
                "id": representative.id,
                "radius": representative.radius,
                "diameter": representative.diameter,
                "layer": representative.layer,
                "building_name": representative.building_name or "Unassigned",
                "building_seq": representative.building_seq,
                "matched_text_height": representative.matched_text.height
                if representative.matched_text
                else "",
                "duplicate_count": len(items),
            }
        )

    error_rows = [
        {
            "error_type": error.error_type,
            "text_id": error.text_id or "",
            "text_value": error.text_value or "",
            "circle_ids": ",".join(error.circle_ids),
            "message": error.message,
        }
        for error in errors
    ]

    summary_map: Dict[str, Dict[str, int]] = defaultdict(
        lambda: {"circles": 0, "texts": 0, "errors": 0}
    )
    for circle in circles:
        name = circle.building_name or "Unassigned"
        summary_map[name]["circles"] += 1
        if circle.has_error:
            summary_map[name]["errors"] += 1
    for text in texts:
        if _is_foundation_pf_text_item(text):
            continue
        name = text.building_name or "Unassigned"
        summary_map[name]["texts"] += 1

    building_order = [b.name for b in buildings] if buildings else []
    if "Unassigned" in summary_map and "Unassigned" not in building_order:
        building_order.append("Unassigned")

    summary_rows = [
        {
            "building": name,
            "circle_count": counts["circles"],
            "text_count": counts["texts"],
            "error_count": counts["errors"],
        }
        for name, counts in summary_map.items()
    ]
    # building 이름으로 오름차순 정렬
    summary_rows.sort(key=lambda x: x["building"] or "")

    return rows, error_rows, summary_rows, building_order or list(summary_map.keys())


def _build_export_stream(
    rows: List[Dict],
    error_rows: List[Dict],
    summary_rows: List[Dict],
    building_order: List[str],
    format: str,
) -> Response:
    """공통 export: rows/error_rows/summary_rows/building_order로 CSV 또는 XLSX 본문 생성."""
    # Summary 다음 탭 순서: 동 이름 오름차순
    building_order = sorted(building_order or ["Unassigned"], key=lambda x: (x or "").strip())
    column_order = [
        PILE_NO_HEADER,
        PILE_Y_HEADER,
        PILE_X_HEADER,
        "id",
        "radius",
        "diameter",
        "layer",
        "building_name",
        "building_seq",
        "matched_text_height",
        "duplicate_count",
    ]
    df = pd.DataFrame(rows)
    existing_columns = [col for col in column_order if col in df.columns]
    remaining_columns = [col for col in df.columns if col not in existing_columns]
    df = df[existing_columns + remaining_columns]

    def parse_pile_no(value):
        if not value or value == "":
            return float("inf")
        try:
            return float(value)
        except (ValueError, TypeError):
            return float("inf")

    if PILE_NO_HEADER in df.columns:
        df["_pile_no_sort"] = df[PILE_NO_HEADER].apply(parse_pile_no)
        df = df.sort_values("_pile_no_sort", na_position="last")
        df = df.drop(columns=["_pile_no_sort"])
    errors_df = pd.DataFrame(error_rows)
    summary_df = pd.DataFrame(summary_rows)

    if format.lower() == "csv":
        buffer = io.StringIO()
        buffer.write("\ufeff")
        if not df.empty:
            written = False
            for building_name in building_order or ["Unassigned"]:
                filtered = df[df["building_name"] == building_name]
                if filtered.empty:
                    continue
                buffer.write(f"\n=== {building_name or '기타'} ===\n")
                filtered_export = filtered.drop(columns=["building_name"], errors="ignore")
                export_column_order = [col for col in column_order if col != "building_name"]
                export_columns = [col for col in export_column_order if col in filtered_export.columns]
                export_columns += [col for col in filtered_export.columns if col not in export_columns]
                filtered_export = filtered_export[export_columns]
                buffer.write(filtered_export.to_csv(index=False))
                written = True
            unassigned_mask = (
                df["building_name"].isna()
                | (df["building_name"] == "Unassigned")
                | (df["building_name"] == "미할당")
            )
            unassigned = df[unassigned_mask]
            if not unassigned.empty:
                unassigned_names = set(unassigned["building_name"].dropna().unique())
                if not any(name in (building_order or []) for name in unassigned_names if name):
                    buffer.write("\n=== 기타 ===\n")
                    unassigned_export = unassigned.drop(columns=["building_name"], errors="ignore")
                    export_column_order = [col for col in column_order if col != "building_name"]
                    export_columns = [col for col in export_column_order if col in unassigned_export.columns]
                    export_columns += [col for col in unassigned_export.columns if col not in export_columns]
                    unassigned_export = unassigned_export[export_columns]
                    buffer.write(unassigned_export.to_csv(index=False))
                    written = True
            if not written:
                df_export = df.drop(columns=["building_name"], errors="ignore")
                export_column_order = [col for col in column_order if col != "building_name"]
                export_columns = [col for col in export_column_order if col in df_export.columns]
                export_columns += [col for col in df_export.columns if col not in export_columns]
                df_export = df_export[export_columns]
                buffer.write(df_export.to_csv(index=False))
        else:
            buffer.write("No clean rows available\n")
        if not errors_df.empty:
            buffer.write("\n\n=== Errors ===\n")
            buffer.write(errors_df.to_csv(index=False))
        buffer.seek(0)
        # StreamingResponse 는 Win/py3.12 + 동시 요청 시 빈 chunk 로 asyncio assert 가 날 수 있음
        return Response(
            content=buffer.getvalue().encode("utf-8"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": 'attachment; filename="circles.csv"'},
        )

    binary = io.BytesIO()
    # 표는 3행부터(1·2행: 제목·여백). pandas startrow=2 → 헤더가 엑셀 3행.
    _xlsx_table_startrow = 2
    _xlsx_freeze = (3, 0)  # A4 기준 고정: 제목·여백·헤더 유지
    titles_by_sheet: Dict[str, str] = {}

    def _export_title_for_building(name: object) -> str:
        s = (str(name).strip() if name is not None else "") or "기타"
        if s in ("Unassigned", "미할당"):
            return "미할당"
        return s

    with pd.ExcelWriter(binary, engine="openpyxl") as writer:
        summary_sheet = (
            summary_df
            if not summary_df.empty
            else pd.DataFrame(columns=["building", "circle_count", "text_count", "error_count"])
        )
        titles_by_sheet["Summary"] = "동별 요약"
        summary_sheet.to_excel(
            writer,
            sheet_name="Summary",
            index=False,
            startrow=_xlsx_table_startrow,
            freeze_panes=_xlsx_freeze,
        )
        if not df.empty:
            written = False
            for building_name in building_order or ["Unassigned"]:
                filtered = df[df["building_name"] == building_name]
                if filtered.empty:
                    continue
                sheet_name = (building_name or "기타")[:31] or "기타"
                titles_by_sheet[sheet_name] = _export_title_for_building(building_name)
                filtered_export = filtered.drop(columns=["building_name"], errors="ignore")
                export_column_order = [col for col in column_order if col != "building_name"]
                export_columns = [col for col in export_column_order if col in filtered_export.columns]
                export_columns += [col for col in filtered_export.columns if col not in export_columns]
                filtered_export = filtered_export[export_columns]
                if PILE_NO_HEADER in filtered_export.columns:
                    filtered_export["_pile_no_sort"] = filtered_export[PILE_NO_HEADER].apply(parse_pile_no)
                    filtered_export = filtered_export.sort_values("_pile_no_sort", na_position="last")
                    filtered_export = filtered_export.drop(columns=["_pile_no_sort"])
                filtered_export.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                    startrow=_xlsx_table_startrow,
                    freeze_panes=_xlsx_freeze,
                )
                written = True
            unassigned_mask = (
                df["building_name"].isna()
                | (df["building_name"] == "Unassigned")
                | (df["building_name"] == "미할당")
            )
            unassigned = df[unassigned_mask]
            if not unassigned.empty:
                unassigned_names = set(unassigned["building_name"].dropna().unique())
                if not any(name in (building_order or []) for name in unassigned_names if name):
                    unassigned_export = unassigned.drop(columns=["building_name"], errors="ignore")
                    export_column_order = [col for col in column_order if col != "building_name"]
                    export_columns = [col for col in export_column_order if col in unassigned_export.columns]
                    export_columns += [col for col in unassigned_export.columns if col not in export_columns]
                    unassigned_export = unassigned_export[export_columns]
                    if PILE_NO_HEADER in unassigned_export.columns:
                        unassigned_export["_pile_no_sort"] = unassigned_export[PILE_NO_HEADER].apply(parse_pile_no)
                        unassigned_export = unassigned_export.sort_values("_pile_no_sort", na_position="last")
                        unassigned_export = unassigned_export.drop(columns=["_pile_no_sort"])
                    titles_by_sheet["기타"] = "기타"
                    unassigned_export.to_excel(
                        writer,
                        sheet_name="기타",
                        index=False,
                        startrow=_xlsx_table_startrow,
                        freeze_panes=_xlsx_freeze,
                    )
                    written = True
            if not written:
                df_export = df.drop(columns=["building_name"], errors="ignore")
                export_column_order = [col for col in column_order if col != "building_name"]
                export_columns = [col for col in export_column_order if col in df_export.columns]
                export_columns += [col for col in df_export.columns if col not in export_columns]
                df_export = df_export[export_columns]
                if PILE_NO_HEADER in df_export.columns:
                    df_export["_pile_no_sort"] = df_export[PILE_NO_HEADER].apply(parse_pile_no)
                    df_export = df_export.sort_values("_pile_no_sort", na_position="last")
                    df_export = df_export.drop(columns=["_pile_no_sort"])
                titles_by_sheet["전체"] = "전체"
                df_export.to_excel(
                    writer,
                    sheet_name="전체",
                    index=False,
                    startrow=_xlsx_table_startrow,
                    freeze_panes=_xlsx_freeze,
                )
        else:
            df_export = df.drop(columns=["building_name"], errors="ignore")
            export_column_order = [col for col in column_order if col != "building_name"]
            export_columns = [col for col in export_column_order if col in df_export.columns]
            export_columns += [col for col in df_export.columns if col not in export_columns]
            df_export = df_export[export_columns]
            if PILE_NO_HEADER in df_export.columns:
                df_export["_pile_no_sort"] = df_export[PILE_NO_HEADER].apply(parse_pile_no)
                df_export = df_export.sort_values("_pile_no_sort", na_position="last")
                df_export = df_export.drop(columns=["_pile_no_sort"])
            titles_by_sheet["Circles"] = "전체"
            df_export.to_excel(
                writer,
                sheet_name="Circles",
                index=False,
                startrow=_xlsx_table_startrow,
                freeze_panes=_xlsx_freeze,
            )
        if not errors_df.empty:
            titles_by_sheet["Errors"] = "오류 내역"
            errors_df.to_excel(
                writer,
                sheet_name="Errors",
                index=False,
                startrow=_xlsx_table_startrow,
                freeze_panes=_xlsx_freeze,
            )
        workbook = writer.book
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        _header_fill = PatternFill(
            start_color="E8EEF5", end_color="E8EEF5", fill_type="solid"
        )
        _thin_black = Side(style="thin", color="000000")
        _all_thin_border = Border(
            left=_thin_black, right=_thin_black, top=_thin_black, bottom=_thin_black
        )
        _title_font = Font(size=16, bold=True)

        def _cell_display_width_units(val: object) -> float:
            if val is None:
                return 0.0
            return float(
                sum(2 if ord(c) > 127 else 1 for c in str(val))
            )

        hdr_row = _xlsx_table_startrow + 1  # 엑셀에서 표 헤더 행 번호

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            max_row, max_col = sheet.max_row, sheet.max_column
            if max_row < 1 or max_col < 1:
                continue
            title_label = titles_by_sheet.get(sheet_name) or sheet_name
            sheet.merge_cells(
                start_row=1, start_column=1, end_row=1, end_column=max_col
            )
            tcell = sheet.cell(row=1, column=1)
            tcell.value = title_label
            tcell.font = _title_font
            tcell.alignment = Alignment(horizontal="left", vertical="center")
            sheet.row_dimensions[1].height = 26

            if max_row >= hdr_row:
                for row in sheet.iter_rows(
                    min_row=hdr_row, max_row=max_row, min_col=1, max_col=max_col
                ):
                    for cell in row:
                        cell.border = _all_thin_border
                header_row_cells = sheet[hdr_row]
                for cell in header_row_cells:
                    cell.fill = _header_fill
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                widest = 0.0
                for row_idx in range(hdr_row, max_row + 1):
                    widest = max(
                        widest,
                        _cell_display_width_units(
                            sheet.cell(row=row_idx, column=col_idx).value
                        ),
                    )
                sheet.column_dimensions[col_letter].width = min(
                    max(widest + 2.8, 9), 72
                )
            y_col_idx = x_col_idx = None
            for col_idx, cell in enumerate(sheet[hdr_row], start=1):
                if cell.value == PILE_Y_HEADER:
                    y_col_idx = col_idx
                elif cell.value == PILE_X_HEADER:
                    x_col_idx = col_idx
            if y_col_idx:
                y_col_letter = get_column_letter(y_col_idx)
                prev = sheet.column_dimensions[y_col_letter].width or 9
                sheet.column_dimensions[y_col_letter].width = max(prev, 12)
                for row in sheet.iter_rows(
                    min_row=hdr_row + 1,
                    max_row=sheet.max_row,
                    min_col=y_col_idx,
                    max_col=y_col_idx,
                ):
                    for cell in row:
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = "0.000"
            if x_col_idx:
                x_col_letter = get_column_letter(x_col_idx)
                prev = sheet.column_dimensions[x_col_letter].width or 9
                sheet.column_dimensions[x_col_letter].width = max(prev, 12)
                for row in sheet.iter_rows(
                    min_row=hdr_row + 1,
                    max_row=sheet.max_row,
                    min_col=x_col_idx,
                    max_col=x_col_idx,
                ):
                    for cell in row:
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = "0.000"
    binary.seek(0)
    return Response(
        content=binary.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="circles.xlsx"'},
    )


def _log_upload_failure(temp_path: Optional[str]) -> None:
    """Log first bytes of temp file when DXF parse fails (diagnostic)."""
    if not temp_path or not os.path.isfile(temp_path):
        return
    try:
        with open(temp_path, "rb") as f:
            head = f.read(200)
        if head:
            logger.warning(
                "DXF parse failed; temp file head (%d bytes): %s",
                len(head),
                head.hex() if len(head) <= 200 else head[:200].hex() + "...",
            )
    except Exception as e:
        logger.debug("Could not read temp file for diagnostic: %s", e)


@app.get("/")
async def root():
    if os.path.exists(FRONTEND_INDEX_PATH):
        return FileResponse(FRONTEND_INDEX_PATH)
    return {"message": "DXF Circle/Text API is running. Open /docs or serve frontend/index.html."}


@app.post("/api/upload-dxf", response_model=CircleResponse)
async def upload_dxf(
    request: Request,
    min_diameter: float = Query(DEFAULT_MIN_DIAMETER),
    max_diameter: float = Query(DEFAULT_MAX_DIAMETER),
    min_area: Optional[float] = Query(None),
    max_area: Optional[float] = Query(None),
    text_height_min: float = Query(DEFAULT_TEXT_HEIGHT_MIN),
    text_height_max: float = Query(DEFAULT_TEXT_HEIGHT_MAX),
    max_match_distance: float = Query(DEFAULT_MAX_MATCH_DISTANCE),
    expected_buildings: Optional[int] = Query(None),
    project_context: Optional[str] = Query(None),
    source_type: Optional[str] = Query(None),
    filename: Optional[str] = Query(None),
    reuse_manual_history: bool = Query(True),
    manual_history_reference_source: Optional[str] = Query(
        None,
        description="(레거시) contractor_original / vendor_prepared. work_id가 있으면 무시.",
    ),
    manual_history_reference_work_id: Optional[str] = Query(
        None,
        description="저장 작업 ID. 지정 시 해당 버전의 수동 매칭만 재사용.",
    ),
) -> CircleResponse:
    if not filename or not filename.lower().endswith(".dxf"):
        raise HTTPException(status_code=400, detail="Only DXF files are supported (query param: filename=...).")

    validate_range("Diameter", min_diameter, max_diameter)
    if min_area is not None and max_area is not None:
        validate_range("Area", min_area, max_area)
    validate_range("Text height", text_height_min, text_height_max)
    if max_match_distance <= 0:
        raise HTTPException(
            status_code=400, detail="Max match distance must be greater than zero."
        )
    temp_path = None
    try:
        # 스트리밍으로 본문 수신: 서버가 청크 단위로 읽어야 클라이언트 진행률이 올라감.
        # (Starlette/FastAPI UploadFile.read()는 전체 업로드 완료까지 대기하는 동작이라 진행률이 1%에서 멈춤)
        # N청크마다만 yield해 버퍼를 더 빨리 소비. 프록시(nginx) 사용 시 proxy_request_buffering off 권장.
        YIELD_EVERY_CHUNKS = 32
        with tempfile.NamedTemporaryFile(delete=False, suffix=".dxf") as tmp:
            temp_path = tmp.name
            total = 0
            chunk_count = 0
            async for chunk in request.stream():
                if chunk:
                    tmp.write(chunk)
                    total += len(chunk)
                    chunk_count += 1
                    if chunk_count % YIELD_EVERY_CHUNKS == 0:
                        await asyncio.sleep(0)
                else:
                    await asyncio.sleep(0)
            tmp.flush()
        if total == 0:
            if temp_path and os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except Exception:
                    pass
            temp_path = None
            raise HTTPException(
                status_code=400,
                detail="업로드된 파일이 비어 있습니다. DXF 파일을 선택했는지 확인하세요.",
            )
        logger.info("upload temp size: %d", total)

        # DXF 파싱은 CPU 부하가 크므로 스레드 풀에서 실행.
        loop = asyncio.get_event_loop()
        try:
            circles, texts, polylines = await loop.run_in_executor(
                None,
                lambda: parse_dxf_entities(
                    temp_path, text_height_min, text_height_max
                ),
            )
        except (ezdxf.DXFStructureError, ezdxf.DXFError) as dxf_err:
            msg = str(dxf_err) if dxf_err else "DXF 파일 형식이 올바르지 않습니다."
            if "is not a DXF file" in msg or "not a DXF" in msg.lower():
                _log_upload_failure(temp_path)
                size_hint = f" (수신: {total} bytes)"
                raise HTTPException(
                    status_code=400,
                    detail="선택한 파일이 DXF 형식이 아니거나 손상되었습니다. CAD에서 '다른 이름으로 저장' → DXF(R12/LT2 또는 최신 ASCII)로 다시 저장한 뒤 시도해 보세요." + size_hint,
                ) from dxf_err
            raise HTTPException(status_code=400, detail=f"DXF 파싱 오류: {msg}") from dxf_err
        except Exception as parse_err:
            msg = (str(parse_err) or "").lower()
            if "is not a dxf file" in msg or "not a dxf" in msg:
                _log_upload_failure(temp_path)
                size_hint = f" (수신: {total} bytes)"
                raise HTTPException(
                    status_code=400,
                    detail="선택한 파일이 DXF 형식이 아니거나 손상되었습니다. CAD에서 '다른 이름으로 저장' → DXF(R12/LT2 또는 최신 ASCII)로 다시 저장한 뒤 시도해 보세요." + size_hint,
                ) from parse_err
            raise
        logger.info(f"Parsed DXF: {len(circles) if circles else 0} circles, {len(texts) if texts else 0} texts, {len(polylines) if polylines else 0} polylines")
        global last_raw_circles, last_raw_texts, last_raw_polylines, building_definitions, last_dxf_file_path
        # 빈 리스트도 허용 (None이 아닌 이상)
        last_raw_circles = circles if circles is not None else []
        last_raw_texts = texts if texts is not None else []
        last_raw_polylines = polylines if polylines is not None else []
        building_definitions = []
        set_expected_building_clusters(expected_buildings)
        hist_ref = (manual_history_reference_source or "").strip() or None
        hist_work = (manual_history_reference_work_id or "").strip() or None
        refresh_manual_history_overrides(
            project_context,
            source_type,
            reuse=reuse_manual_history,
            history_reference_source_type=hist_ref,
            history_reference_work_id=hist_work,
            clear_session_overrides=True,
        )

        # DXF 파일 경로 저장 (텍스트 높이 필터 재적용용)
        # 기존 파일이 있으면 삭제
        if last_dxf_file_path and os.path.exists(last_dxf_file_path):
            try:
                os.remove(last_dxf_file_path)
            except Exception:
                pass
        # 새 파일을 영구적으로 저장 (재파싱용)
        import uuid as uuid_module
        saved_file_path = os.path.join(tempfile.gettempdir(), f"dxf_upload_{uuid_module.uuid4().hex}.dxf")
        shutil.copy2(temp_path, saved_file_path)
        last_dxf_file_path = saved_file_path

        filters = FilterSettings(
            min_diameter=min_diameter,
            max_diameter=max_diameter,
            min_area=min_area,
            max_area=max_area,
            text_height_min=text_height_min,
            text_height_max=text_height_max,
            max_match_distance=max_match_distance,
        )
        # 업로드 직후이므로 전역 변수가 설정되었는지 확인
        if last_raw_circles is None or last_raw_texts is None:
            raise HTTPException(status_code=500, detail="Failed to parse DXF file data.")
        return update_cached_result(filters)
    except HTTPException:
        raise
    except (ezdxf.DXFStructureError, ezdxf.DXFError) as exc:
        raise HTTPException(
            status_code=400, detail=f"Unable to parse DXF file: {exc}"
        ) from exc
    except Exception as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    finally:
        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)


# Cloudflare Tunnel 등에서 대용량 본문이 1~2MB만 전달된 뒤 끊길 때 사용. 파일을 잘라 여러 요청으로 보냄.
CHUNKED_UPLOAD_CHUNK_SIZE = 1024 * 1024  # 1 MB
CHUNKED_UPLOAD_DIR = os.path.join(tempfile.gettempdir(), "pilexy_upload_chunks")


def _chunked_upload_dir(upload_id: str) -> str:
    safe_id = re.sub(r"[^\w\-]", "", upload_id)[:64]
    d = os.path.join(CHUNKED_UPLOAD_DIR, safe_id)
    os.makedirs(d, exist_ok=True)
    return d


@app.post("/api/upload-dxf-chunk")
async def upload_dxf_chunk(
    request: Request,
    upload_id: str = Query(..., description="동일 업로드 세션 식별자"),
    chunk_index: int = Query(..., ge=0),
    total_chunks: int = Query(..., ge=1),
    filename: str = Query(...),
    min_diameter: float = Query(DEFAULT_MIN_DIAMETER),
    max_diameter: float = Query(DEFAULT_MAX_DIAMETER),
    min_area: Optional[float] = Query(None),
    max_area: Optional[float] = Query(None),
    text_height_min: float = Query(DEFAULT_TEXT_HEIGHT_MIN),
    text_height_max: float = Query(DEFAULT_TEXT_HEIGHT_MAX),
    max_match_distance: float = Query(DEFAULT_MAX_MATCH_DISTANCE),
    expected_buildings: Optional[int] = Query(None),
    project_context: Optional[str] = Query(None),
    source_type: Optional[str] = Query(None),
    reuse_manual_history: bool = Query(True),
    manual_history_reference_source: Optional[str] = Query(
        None,
        description="(레거시) 비우면 source_type과 동일한 히스토리 버킷.",
    ),
    manual_history_reference_work_id: Optional[str] = Query(
        None,
        description="저장 작업 ID. 지정 시 해당 버전의 수동 매칭만 재사용.",
    ),
):
    """대용량 DXF를 1MB 단위로 잘라 보낼 때 사용. 마지막 청크 수신 시 병합 후 파싱해 CircleResponse 반환."""
    if not filename.lower().endswith(".dxf"):
        raise HTTPException(status_code=400, detail="Only DXF files are supported.")
    validate_range("Diameter", min_diameter, max_diameter)
    if min_area is not None and max_area is not None:
        validate_range("Area", min_area, max_area)
    validate_range("Text height", text_height_min, text_height_max)
    if max_match_distance <= 0:
        raise HTTPException(status_code=400, detail="Max match distance must be greater than zero.")
    body = await request.body()
    chunk_dir = _chunked_upload_dir(upload_id)
    chunk_path = os.path.join(chunk_dir, f"{chunk_index}.bin")
    with open(chunk_path, "wb") as f:
        f.write(body)
    if chunk_index < total_chunks - 1:
        return {"ok": True, "chunk": chunk_index + 1, "total_chunks": total_chunks}
    # 마지막 청크: 병합 후 파싱
    merge_path = None
    try:
        merge_path = os.path.join(tempfile.gettempdir(), f"dxf_merge_{uuid.uuid4().hex}.dxf")
        with open(merge_path, "wb") as out:
            for i in range(total_chunks):
                p = os.path.join(chunk_dir, f"{i}.bin")
                if not os.path.isfile(p):
                    raise HTTPException(status_code=400, detail=f"Missing chunk {i}.")
                with open(p, "rb") as f:
                    out.write(f.read())
        for i in range(total_chunks):
            try:
                os.remove(os.path.join(chunk_dir, f"{i}.bin"))
            except Exception:
                pass
        try:
            os.rmdir(chunk_dir)
        except Exception:
            pass
        loop = asyncio.get_event_loop()
        circles, texts, polylines = await loop.run_in_executor(
            None,
            lambda: parse_dxf_entities(merge_path, text_height_min, text_height_max),
        )
        global last_raw_circles, last_raw_texts, last_raw_polylines, building_definitions, last_dxf_file_path
        last_raw_circles = circles if circles is not None else []
        last_raw_texts = texts if texts is not None else []
        last_raw_polylines = polylines if polylines is not None else []
        building_definitions = []
        set_expected_building_clusters(expected_buildings)
        hist_ref = (manual_history_reference_source or "").strip() or None
        hist_work = (manual_history_reference_work_id or "").strip() or None
        refresh_manual_history_overrides(
            project_context,
            source_type,
            reuse=reuse_manual_history,
            history_reference_source_type=hist_ref,
            history_reference_work_id=hist_work,
            clear_session_overrides=True,
        )
        if last_dxf_file_path and os.path.exists(last_dxf_file_path):
            try:
                os.remove(last_dxf_file_path)
            except Exception:
                pass
        saved_file_path = os.path.join(tempfile.gettempdir(), f"dxf_upload_{uuid.uuid4().hex}.dxf")
        shutil.copy2(merge_path, saved_file_path)
        last_dxf_file_path = saved_file_path
        filters = FilterSettings(
            min_diameter=min_diameter,
            max_diameter=max_diameter,
            min_area=min_area,
            max_area=max_area,
            text_height_min=text_height_min,
            text_height_max=text_height_max,
            max_match_distance=max_match_distance,
        )
        return update_cached_result(filters)
    except HTTPException:
        raise
    except (ezdxf.DXFStructureError, ezdxf.DXFError) as e:
        raise HTTPException(status_code=400, detail=f"DXF parse error: {e}") from e
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e
    finally:
        if merge_path and os.path.exists(merge_path):
            try:
                os.remove(merge_path)
            except Exception:
                pass


@app.get("/api/circles", response_model=CircleResponse)
async def get_circles(
    min_diameter: Optional[float] = Query(None),
    max_diameter: Optional[float] = Query(None),
    min_area: Optional[float] = Query(None),
    max_area: Optional[float] = Query(None),
    text_height_min: Optional[float] = Query(None),
    text_height_max: Optional[float] = Query(None),
    max_match_distance: Optional[float] = Query(None),
    text_reference_point: Optional[str] = Query(None),
    expected_buildings: Optional[int] = Query(None),
    max_distance_from_seed: Optional[float] = Query(None),
    merge_seed_distance: Optional[float] = Query(None),
) -> CircleResponse:
    global clustering_max_distance_from_seed, clustering_merge_seed_distance, last_raw_circles, last_raw_texts, last_raw_polylines, last_dxf_file_path
    
    filters = resolve_filter_values(
        min_diameter, max_diameter, min_area, max_area,
        text_height_min, text_height_max, max_match_distance, text_reference_point
    )
    
    # 텍스트 높이 필터가 변경되었는지 확인
    text_height_changed = False
    if current_filters and (last_dxf_file_path and os.path.exists(last_dxf_file_path)):
        # 텍스트 높이 필터가 실제로 변경되었는지 확인
        if (filters.text_height_min != current_filters.text_height_min or 
            filters.text_height_max != current_filters.text_height_max):
            text_height_changed = True
            logger.info(f"텍스트 높이 필터 변경 감지: {current_filters.text_height_min}-{current_filters.text_height_max} -> {filters.text_height_min}-{filters.text_height_max}")
    
    # 텍스트 높이 필터가 변경되었으면 DXF 파일을 다시 파싱 (스레드 풀에서 실행해 이벤트 루프 블로킹 방지)
    if text_height_changed and last_dxf_file_path:
        try:
            logger.info(f"텍스트 높이 필터 변경으로 인한 DXF 재파싱: {last_dxf_file_path}")
            loop = asyncio.get_event_loop()
            circles, texts, polylines = await loop.run_in_executor(
                None,
                lambda: parse_dxf_entities(
                    last_dxf_file_path, filters.text_height_min, filters.text_height_max
                ),
            )
            last_raw_circles = circles if circles is not None else []
            last_raw_texts = texts if texts is not None else []
            last_raw_polylines = polylines if polylines is not None else []
            logger.info(f"재파싱 완료: {len(last_raw_circles)} circles, {len(last_raw_texts)} texts")
        except Exception as exc:
            logger.error(f"DXF 재파싱 실패: {exc}")
            # 재파싱 실패 시 기존 데이터 사용
    
    if expected_buildings is not None:
        set_expected_building_clusters(expected_buildings)
    if max_distance_from_seed is not None:
        clustering_max_distance_from_seed = max_distance_from_seed
    if merge_seed_distance is not None:
        clustering_merge_seed_distance = merge_seed_distance
    return update_cached_result(filters)


@app.post("/api/circles/refresh-manual-history-match", response_model=CircleResponse)
async def refresh_manual_history_match(payload: ManualHistoryMatchRefreshRequest) -> CircleResponse:
    """히스토리 재적용 여부·참고 구분을 바꾸고 현재 필터로 다시 매칭. DXF가 없으면 본문의 circles/texts로 서버 원시 데이터를 채운다."""
    global last_raw_circles, last_raw_texts, last_raw_polylines, building_definitions, manual_overrides
    global history_manual_overrides, current_filters, clustering_max_distance_from_seed
    global clustering_merge_seed_distance, last_dxf_file_path

    if payload.circles and len(payload.circles) > 0:
        last_raw_circles = [_normalize_circle_for_filter(c) for c in payload.circles]
        last_raw_texts = [_normalize_text_for_filter(t) for t in (payload.texts or [])]
        if payload.polylines is not None:
            last_raw_polylines = copy.deepcopy(payload.polylines) if payload.polylines else []
        if payload.buildings is not None:
            building_definitions = list(payload.buildings)
        last_dxf_file_path = None
        mo: Dict[str, str] = {}
        if payload.manual_overrides is not None:
            for k, v in payload.manual_overrides.items():
                if k is not None and v is not None:
                    sk, sv = str(k), str(v)
                    if sk and sv:
                        mo[sk] = sv
            for c in payload.circles:
                if not isinstance(c, dict):
                    continue
                cid = c.get("id")
                if cid is None or not c.get("manual_match") or not c.get("matched_text_id"):
                    continue
                mo[str(cid)] = str(c["matched_text_id"])
            manual_overrides = mo
        else:
            manual_overrides = {}
            for c in payload.circles:
                if not isinstance(c, dict):
                    continue
                cid = c.get("id")
                if cid is None or not c.get("manual_match") or not c.get("matched_text_id"):
                    continue
                manual_overrides[str(cid)] = str(c["matched_text_id"])
        if payload.filter is not None:
            current_filters = payload.filter
        if payload.expected_buildings is not None:
            set_expected_building_clusters(payload.expected_buildings)
        if payload.max_distance_from_seed is not None:
            clustering_max_distance_from_seed = payload.max_distance_from_seed
        if payload.merge_seed_distance is not None:
            clustering_merge_seed_distance = payload.merge_seed_distance

    if last_raw_circles is None or last_raw_texts is None or not last_raw_circles or not last_raw_texts:
        raise HTTPException(
            status_code=400,
            detail="도면 데이터가 없습니다. 화면에 circle/text가 있어야 하며, 저장 작업 불러오기 후에는 요청에 circles/texts가 포함되어야 합니다.",
        )
    project = normalize_project_name(payload.project_context) if payload.project_context else DEFAULT_PROJECT_NAME
    work_st = normalize_source_type(payload.source_type) if payload.source_type else DEFAULT_SOURCE_TYPE
    hist_ref_raw = (payload.history_reference_source_type or "").strip() or None
    hist_work_raw = (payload.history_reference_work_id or "").strip() or None
    refresh_manual_history_overrides(
        project,
        work_st,
        reuse=payload.reuse_manual_history,
        history_reference_source_type=hist_ref_raw,
        history_reference_work_id=hist_work_raw,
        clear_session_overrides=False,
    )
    filters = current_filters or FilterSettings(
        min_diameter=DEFAULT_MIN_DIAMETER,
        max_diameter=DEFAULT_MAX_DIAMETER,
        text_height_min=DEFAULT_TEXT_HEIGHT_MIN,
        text_height_max=DEFAULT_TEXT_HEIGHT_MAX,
        max_match_distance=DEFAULT_MAX_MATCH_DISTANCE,
        text_reference_point="center",
    )
    return update_cached_result(filters)


@app.post("/api/circles/recompute", response_model=CircleResponse)
async def recompute_from_circles(payload: RecomputeRequest) -> CircleResponse:
    """
    불러온 circles로 클러스터/동 윤곽선만 재계산. (저장 데이터 불러오기 후 윤곽선 다시 생성·동 개수 적용용)
    """
    if not payload.circles or len(payload.circles) == 0:
        raise HTTPException(status_code=400, detail="circles is required and must not be empty.")
    circles = copy.deepcopy(payload.circles)
    texts = copy.deepcopy(payload.texts) if payload.texts else []
    for c in circles:
        if "id" not in c or "center_x" not in c or "center_y" not in c:
            raise HTTPException(
                status_code=400,
                detail="Each circle must have id, center_x, center_y.",
            )
        c.setdefault("center_z", 0.0)
        c.setdefault("radius", 0.0)
        c.setdefault("diameter", 0.0)
        c.setdefault("layer", "0")
    expected = max(1, int(payload.expected_buildings))
    max_dist = payload.max_distance_from_seed if payload.max_distance_from_seed is not None else 30.0
    merge_dist = payload.merge_seed_distance if payload.merge_seed_distance is not None else 20.0

    pile_points = build_pile_points_for_clustering(circles)
    try:
        cluster_results = cluster_piles(
            pile_points,
            DEFAULT_CLUSTER_CONFIG,
            expected_clusters=expected,
            max_distance_from_seed=max_dist,
            merge_seed_distance=merge_dist,
        )
    except Exception as e:
        logger.exception("Recompute cluster_piles failed: %s", e)
        raise HTTPException(status_code=500, detail=str(e)) from e

    cluster_polylines = clusters_to_polylines(cluster_results)
    buildings: List[BuildingDefinition] = []
    for i, poly in enumerate(cluster_polylines):
        points = poly.get("points") or []
        verts = [BuildingVertex(x=float(p.get("x", 0)), y=float(p.get("y", 0))) for p in points]
        if len(verts) >= 3:
            name = poly.get("cluster_id") or poly.get("id") or f"building_{i + 1}"
            buildings.append(BuildingDefinition(name=name, kind="building", vertices=verts))

    classify_entities(circles, texts, buildings, cluster_polylines if cluster_polylines else None)
    building_summary = compute_building_summary(circles, texts, buildings)
    matched = sum(1 for c in circles if c.get("matched_text_id"))
    summary = SummaryStats(
        total_circles=len(circles),
        total_texts=_count_non_foundation_texts(texts),
        matched_pairs=matched,
        duplicate_groups=0,
    )
    filters = current_filters or FilterSettings(
        min_diameter=DEFAULT_MIN_DIAMETER,
        max_diameter=DEFAULT_MAX_DIAMETER,
        text_height_min=DEFAULT_TEXT_HEIGHT_MIN,
        text_height_max=DEFAULT_TEXT_HEIGHT_MAX,
        max_match_distance=DEFAULT_MAX_MATCH_DISTANCE,
        text_reference_point="center",
    )
    circle_records = [to_circle_model(c) for c in circles]
    text_records = [to_text_model(t) for t in texts] if texts else []
    cluster_payload = [build_cluster_payload(cluster) for cluster in cluster_results]
    return CircleResponse(
        summary=summary,
        circles=circle_records,
        texts=text_records,
        duplicates=[],
        polylines=[],
        buildings=buildings,
        pile_clusters=cluster_payload,
        cluster_polylines=_ensure_polyline_ids(cluster_polylines),
        errors=[],
        filter=filters,
        building_summary=building_summary,
        match_corrections=[to_match_correction_model(item) for item in last_match_corrections],
    )


def _ensure_polyline_ids(polylines: List[Dict]) -> List[Dict]:
    """CircleResponse 검증 통과용: id가 없는 폴리라인에 poly_0, poly_1, ... 부여."""
    if not polylines:
        return []
    out = []
    for i, p in enumerate(polylines):
        if not isinstance(p, dict):
            continue
        d = dict(p)
        d.setdefault("id", p.get("id") or f"poly_{i}")
        d.setdefault("closed", True)
        d.setdefault("points", [])
        out.append(d)
    return out


def _normalize_circle_for_filter(c: Dict) -> Dict:
    """Apply-filter 요청용 circle dict 정규화 (radius/diameter/area, snake_case 등)."""
    out = copy.deepcopy(c) if isinstance(c, dict) else {}
    out.setdefault("id", out.get("id", ""))
    out.setdefault("center_x", float(out.get("center_x", 0)))
    out.setdefault("center_y", float(out.get("center_y", 0)))
    out.setdefault("center_z", float(out.get("center_z", 0)))
    out.setdefault("radius", float(out.get("radius", 0)))
    out.setdefault("diameter", float(out.get("diameter") or (out["radius"] * 2.0)))
    if out.get("area") is None:
        out["area"] = math.pi * out["radius"] * out["radius"]
    out.setdefault("layer", out.get("layer", "0"))
    out.setdefault("block_name", out.get("block_name"))
    return out


def _normalize_text_for_filter(t: Dict) -> Dict:
    """Apply-filter 요청용 text dict 정규화 (text_center_x/y, insert_x/y)."""
    out = copy.deepcopy(t) if isinstance(t, dict) else {}
    out.setdefault("id", out.get("id", ""))
    # API 응답은 center_x/center_y로 올 수 있음
    cx = out.get("text_center_x") or out.get("center_x", 0)
    cy = out.get("text_center_y") or out.get("center_y", 0)
    out["text_center_x"] = float(cx)
    out["text_center_y"] = float(cy)
    out.setdefault("insert_x", float(out.get("insert_x", cx)))
    out.setdefault("insert_y", float(out.get("insert_y", cy)))
    out.setdefault("insert_z", float(out.get("insert_z", 0)))
    out.setdefault("height", float(out.get("height", 0)))
    out.setdefault("text", out.get("text", ""))
    out.setdefault("layer", out.get("layer", "0"))
    out.setdefault("type", out.get("type", "TEXT"))
    out.setdefault("block_name", out.get("block_name"))
    return out


def _build_filtered_circle_response(
    circles: List[Dict[str, Any]],
    texts: List[Dict[str, Any]],
    filters: FilterSettings,
    *,
    expected_buildings: int = 1,
    max_distance_from_seed: Optional[float] = None,
    merge_seed_distance: Optional[float] = None,
    buildings: Optional[List[BuildingDefinition]] = None,
    manual_overrides_map: Optional[Dict[str, str]] = None,
    response_polylines: Optional[List[Dict[str, Any]]] = None,
    match_corrections: Optional[List[MatchCorrection]] = None,
) -> CircleResponse:
    filtered_circles = filter_circles(
        circles, filters.min_diameter, filters.max_diameter, filters.min_area, filters.max_area
    )
    filtered_texts = filter_texts(texts, filters.text_height_min, filters.text_height_max)
    overrides = dict(manual_overrides_map) if manual_overrides_map else {}
    text_lookup_fb = {str(t["id"]): t for t in filtered_texts}
    overrides = _prune_pairwise_swap_manual_overrides(
        filtered_circles,
        text_lookup_fb,
        overrides,
        filters.text_reference_point,
        max_center_separation=max(
            2.5,
            float(filters.max_match_distance) * MANUAL_PRUNE_MAX_CENTER_FACTOR,
        ),
        min_improvement=MANUAL_PRUNE_MIN_DISTANCE_IMPROVEMENT,
    )
    matched_pairs, text_links = match_texts_to_circles(
        filtered_circles,
        filtered_texts,
        filters.max_match_distance,
        overrides=overrides,
        text_reference_point=filters.text_reference_point,
    )
    errors = build_match_errors(filtered_circles, filtered_texts, text_links)
    duplicate_groups = build_duplicate_groups(filtered_circles)

    cluster_payload_list: List[Dict] = []
    cluster_polylines: List[Dict] = []
    if filtered_circles:
        pile_points = build_pile_points_for_clustering(filtered_circles)
        cluster_results = cluster_piles(
            pile_points,
            DEFAULT_CLUSTER_CONFIG,
            expected_clusters=max(1, int(expected_buildings)),
            max_distance_from_seed=max_distance_from_seed if max_distance_from_seed is not None else 30.0,
            merge_seed_distance=merge_seed_distance if merge_seed_distance is not None else 20.0,
        )
        cluster_payload_list = [build_cluster_payload(cluster) for cluster in cluster_results]
        cluster_polylines = clusters_to_polylines(cluster_results)

    resolved_buildings = list(buildings or [])
    if not resolved_buildings:
        for i, poly in enumerate(cluster_polylines):
            points = poly.get("points") or []
            verts = [
                BuildingVertex(x=float(p.get("x", 0)), y=float(p.get("y", 0)))
                for p in points
            ]
            if len(verts) >= 3:
                name = poly.get("cluster_id") or poly.get("id") or f"building_{i + 1}"
                resolved_buildings.append(
                    BuildingDefinition(name=name, kind="building", vertices=verts)
                )

    classify_entities(
        filtered_circles,
        filtered_texts,
        resolved_buildings,
        cluster_polylines if cluster_polylines else None,
    )
    extend_same_building_number_duplicates(filtered_circles, errors, duplicate_groups)
    building_summary = compute_building_summary(
        filtered_circles, filtered_texts, resolved_buildings
    )
    summary = SummaryStats(
        total_circles=len(filtered_circles),
        total_texts=_count_non_foundation_texts(filtered_texts),
        matched_pairs=matched_pairs,
        duplicate_groups=len(duplicate_groups),
    )
    safe_response_polylines = _sanitize_reference_polylines(
        response_polylines, filtered_circles
    )
    circle_records = [to_circle_model(c) for c in filtered_circles]
    text_records = [to_text_model(t) for t in filtered_texts]
    return CircleResponse(
        summary=summary,
        circles=circle_records,
        texts=text_records,
        duplicates=duplicate_groups,
        polylines=_ensure_polyline_ids(safe_response_polylines),
        buildings=resolved_buildings,
        pile_clusters=cluster_payload_list,
        cluster_polylines=_ensure_polyline_ids(cluster_polylines),
        errors=errors,
        filter=filters,
        building_summary=building_summary,
        match_corrections=match_corrections or [],
    )


def _normalize_saved_work_filter(filter_data: Any) -> FilterSettings:
    raw = filter_data if isinstance(filter_data, dict) else {}

    def read_float(*keys: str, default: Optional[float] = None) -> Optional[float]:
        for key in keys:
            value = raw.get(key) if isinstance(raw, dict) else None
            if value is None or value == "":
                continue
            try:
                return float(value)
            except (TypeError, ValueError):
                continue
        return default

    ref_pt = str(
        raw.get("text_reference_point")
        or raw.get("textReferencePoint")
        or "center"
    ).strip().lower()
    if ref_pt not in ("center", "insert"):
        ref_pt = "center"

    min_diameter = read_float("min_diameter", "minDiameter", default=DEFAULT_MIN_DIAMETER)
    max_diameter = read_float("max_diameter", "maxDiameter", default=DEFAULT_MAX_DIAMETER)
    text_height_min = read_float(
        "text_height_min",
        "textHeightMin",
        default=DEFAULT_TEXT_HEIGHT_MIN,
    )
    text_height_max = read_float(
        "text_height_max",
        "textHeightMax",
        default=DEFAULT_TEXT_HEIGHT_MAX,
    )
    max_match_distance = read_float(
        "max_match_distance",
        "maxMatchDistance",
        default=DEFAULT_MAX_MATCH_DISTANCE,
    )
    if max_match_distance is None or max_match_distance <= 0:
        max_match_distance = DEFAULT_MAX_MATCH_DISTANCE

    min_area = read_float("min_area", "minArea")
    max_area = read_float("max_area", "maxArea")
    filters = FilterSettings(
        min_diameter=min_diameter if min_diameter is not None else DEFAULT_MIN_DIAMETER,
        max_diameter=max_diameter if max_diameter is not None else DEFAULT_MAX_DIAMETER,
        min_area=min_area,
        max_area=max_area,
        text_height_min=text_height_min if text_height_min is not None else DEFAULT_TEXT_HEIGHT_MIN,
        text_height_max=text_height_max if text_height_max is not None else DEFAULT_TEXT_HEIGHT_MAX,
        max_match_distance=max_match_distance,
        text_reference_point=ref_pt,
    )

    try:
        validate_range("Diameter", filters.min_diameter, filters.max_diameter)
    except HTTPException:
        filters.min_diameter = DEFAULT_MIN_DIAMETER
        filters.max_diameter = DEFAULT_MAX_DIAMETER
    try:
        validate_range("Text height", filters.text_height_min, filters.text_height_max)
    except HTTPException:
        filters.text_height_min = DEFAULT_TEXT_HEIGHT_MIN
        filters.text_height_max = DEFAULT_TEXT_HEIGHT_MAX
    if filters.min_area is not None and filters.max_area is not None:
        try:
            validate_range("Area", filters.min_area, filters.max_area)
        except HTTPException:
            filters.min_area = None
            filters.max_area = None
    return filters


def _serialize_filter_for_saved_work(filters: FilterSettings) -> Dict[str, Any]:
    payload: Dict[str, Any] = {
        "minDiameter": filters.min_diameter,
        "maxDiameter": filters.max_diameter,
        "textHeightMin": filters.text_height_min,
        "textHeightMax": filters.text_height_max,
        "maxMatchDistance": filters.max_match_distance,
        "textReferencePoint": filters.text_reference_point,
    }
    if filters.min_area is not None:
        payload["minArea"] = filters.min_area
    if filters.max_area is not None:
        payload["maxArea"] = filters.max_area
    return payload


def _merge_client_only_saved_work_filter_fields(
    merged: Dict[str, Any], original_filter: Any
) -> None:
    """저장 작업 재수화 시 FilterSettings에 없는 UI 전용 플래그를 원본 JSON에서 복원."""
    if not isinstance(original_filter, dict):
        return
    for camel, snake in (
        ("pileNumberHyphenFormat", "pile_number_hyphen_format"),
        ("towerCraneNumberFormat", "tower_crane_number_format"),
        ("excludeIdenticalGeometryDuplicates", "exclude_identical_geometry_duplicates"),
    ):
        if camel in original_filter:
            merged[camel] = original_filter[camel]
        elif snake in original_filter:
            merged[camel] = original_filter[snake]


def _normalize_saved_work_buildings(items: Any) -> List[BuildingDefinition]:
    result: List[BuildingDefinition] = []
    if not isinstance(items, list):
        return result
    for item in items:
        try:
            result.append(
                item if isinstance(item, BuildingDefinition) else BuildingDefinition.model_validate(item)
            )
        except Exception:
            continue
    return result


def _rehydrate_saved_work_payload(data: Dict[str, Any]) -> Dict[str, Any]:
    payload = apply_work_defaults(copy.deepcopy(data if isinstance(data, dict) else {}))
    circles_data = payload.get("circles") if isinstance(payload.get("circles"), list) else []
    texts_data = payload.get("texts") if isinstance(payload.get("texts"), list) else []
    if not circles_data:
        return payload

    circles = [_normalize_circle_for_filter(item) for item in circles_data]
    texts = [_normalize_text_for_filter(item) for item in texts_data]
    filters = _normalize_saved_work_filter(payload.get("filter"))
    buildings = _normalize_saved_work_buildings(payload.get("buildings"))
    raw_polylines = copy.deepcopy(
        payload.get("rawPolylines")
        if isinstance(payload.get("rawPolylines"), list)
        else payload.get("polylines")
        if isinstance(payload.get("polylines"), list)
        else []
    )
    manual_overrides_map = {
        str(circle_id): str(text_id)
        for circle_id, text_id in (
            payload.get("manualOverrides")
            if isinstance(payload.get("manualOverrides"), dict)
            else payload.get("manual_overrides")
            if isinstance(payload.get("manual_overrides"), dict)
            else {}
        ).items()
        if circle_id is not None and text_id is not None
    }
    clustering = payload.get("clustering") if isinstance(payload.get("clustering"), dict) else {}
    try:
        expected_buildings = int(
            payload.get("buildingCount")
            or clustering.get("expectedBuildings")
            or len(buildings)
            or 1
        )
    except (TypeError, ValueError):
        expected_buildings = max(1, len(buildings))
    response = _build_filtered_circle_response(
        circles,
        texts,
        filters,
        expected_buildings=max(1, expected_buildings),
        max_distance_from_seed=clustering.get("maxDistanceFromSeed"),
        merge_seed_distance=clustering.get("mergeSeedDistance"),
        buildings=buildings,
        manual_overrides_map=manual_overrides_map,
        response_polylines=raw_polylines,
        match_corrections=[],
    )

    payload["summary"] = response.summary.model_dump()
    payload["circles"] = [item.model_dump() for item in response.circles]
    payload["texts"] = [item.model_dump() for item in response.texts]
    payload["duplicates"] = [item.model_dump() for item in response.duplicates]
    payload["errors"] = [item.model_dump() for item in response.errors]
    _serialized_filter = _serialize_filter_for_saved_work(response.filter)
    _merge_client_only_saved_work_filter_fields(_serialized_filter, payload.get("filter"))
    payload["filter"] = _serialized_filter
    payload["buildings"] = [item.model_dump() for item in response.buildings]
    payload["pileClusters"] = [item.model_dump() for item in response.pile_clusters]
    payload["clusterPolylines"] = [item.model_dump() for item in response.cluster_polylines]
    payload["buildingSummary"] = [item.model_dump() for item in response.building_summary]
    if raw_polylines:
        payload["rawPolylines"] = raw_polylines
        payload["polylines"] = raw_polylines
    payload["matchCorrections"] = (
        payload.get("matchCorrections")
        if isinstance(payload.get("matchCorrections"), list)
        else payload.get("match_corrections")
        if isinstance(payload.get("match_corrections"), list)
        else []
    )
    return payload


@app.post("/api/circles/apply-filter", response_model=CircleResponse)
async def apply_filter_to_circles(payload: ApplyFilterRequest) -> CircleResponse:
    """
    불러온 circles/texts에 필터를 적용하고 매칭·클러스터를 재계산합니다.
    (불러오기 후 필터 수정해서 확인할 때 사용, 동 윤곽선 편집과 동일한 방식.)
    """
    if not payload.circles or len(payload.circles) == 0:
        raise HTTPException(status_code=400, detail="circles is required and must not be empty.")
    circles = [_normalize_circle_for_filter(c) for c in payload.circles]
    texts = [_normalize_text_for_filter(t) for t in (payload.texts or [])]

    min_d = payload.min_diameter if payload.min_diameter is not None else DEFAULT_MIN_DIAMETER
    max_d = payload.max_diameter if payload.max_diameter is not None else DEFAULT_MAX_DIAMETER
    min_a = payload.min_area
    max_a = payload.max_area
    th_min = (
        payload.text_height_min
        if payload.text_height_min is not None
        else DEFAULT_TEXT_HEIGHT_MIN
    )
    th_max = (
        payload.text_height_max
        if payload.text_height_max is not None
        else DEFAULT_TEXT_HEIGHT_MAX
    )
    match_dist = (
        payload.max_match_distance
        if payload.max_match_distance is not None
        else DEFAULT_MAX_MATCH_DISTANCE
    )
    ref_pt = (payload.text_reference_point or "center").lower()
    if ref_pt not in ("center", "insert"):
        ref_pt = "center"
    filters = FilterSettings(
        min_diameter=min_d,
        max_diameter=max_d,
        min_area=min_a,
        max_area=max_a,
        text_height_min=th_min,
        text_height_max=th_max,
        max_match_distance=match_dist,
        text_reference_point=ref_pt,
    )
    validate_range("Diameter", filters.min_diameter, filters.max_diameter)
    if filters.min_area is not None and filters.max_area is not None:
        validate_range("Area", filters.min_area, filters.max_area)
    validate_range("Text height", filters.text_height_min, filters.text_height_max)
    if filters.max_match_distance <= 0:
        raise HTTPException(status_code=400, detail="Max match distance must be greater than zero.")
    try:
        return _build_filtered_circle_response(
            circles,
            texts,
            filters,
            expected_buildings=payload.expected_buildings,
            max_distance_from_seed=payload.max_distance_from_seed,
            merge_seed_distance=payload.merge_seed_distance,
            buildings=list(payload.buildings) if payload.buildings else [],
            manual_overrides_map=payload.manual_overrides,
            response_polylines=[],
            match_corrections=[to_match_correction_model(item) for item in last_match_corrections],
        )
    except Exception as e:
        logger.exception("apply_filter cluster/match pipeline failed: %s", e)
        raise HTTPException(status_code=500, detail=str(e)) from e


@app.post("/api/assign-buildings", response_model=CircleResponse)
async def assign_buildings(payload: BuildingApplyRequest) -> CircleResponse:
    # 클라이언트가 circles를 보내면 해당 데이터 기준으로 적용 (다중 사용자 시 전역 상태 오염 방지)
    if payload.circles and len(payload.circles) > 0:
        global last_raw_circles, last_raw_texts, last_raw_polylines, building_definitions, manual_overrides, history_manual_overrides
        saved_circles = last_raw_circles
        saved_texts = last_raw_texts
        saved_polylines = last_raw_polylines
        saved_buildings = list(building_definitions)
        saved_manual_overrides = dict(manual_overrides)
        saved_history_manual_overrides = dict(history_manual_overrides)
        try:
            # 불러오기/설정만 적용 시 클라이언트 circles 형식 차이로 500 방지 — apply-filter와 동일하게 정규화
            last_raw_circles = [_normalize_circle_for_filter(c) for c in payload.circles]
            last_raw_texts = [_normalize_text_for_filter(t) for t in (payload.texts or [])]
            last_raw_polylines = copy.deepcopy(payload.polylines) if payload.polylines else []
            building_definitions = list(payload.buildings)
            # 수동 매칭: 클라이언트가 보낸 맵으로 서버 전역과 맞춤 (없으면 과거 세션의 stale override가 남아 재매칭이 틀어짐)
            mo: Dict[str, str] = {}
            if payload.manual_overrides is not None:
                for k, v in payload.manual_overrides.items():
                    if k is not None and v is not None:
                        sk, sv = str(k), str(v)
                        if sk and sv:
                            mo[sk] = sv
                for c in payload.circles:
                    if not isinstance(c, dict):
                        continue
                    cid = c.get("id")
                    if cid is None or not c.get("manual_match") or not c.get("matched_text_id"):
                        continue
                    mo[str(cid)] = str(c["matched_text_id"])
                manual_overrides = mo
                history_manual_overrides = {}
            else:
                for c in payload.circles:
                    if not isinstance(c, dict):
                        continue
                    cid = c.get("id")
                    if cid is None or not c.get("manual_match") or not c.get("matched_text_id"):
                        continue
                    manual_overrides[str(cid)] = str(c["matched_text_id"])
                history_manual_overrides = {}
            filters = current_filters or FilterSettings(
                min_diameter=DEFAULT_MIN_DIAMETER,
                max_diameter=DEFAULT_MAX_DIAMETER,
                text_height_min=DEFAULT_TEXT_HEIGHT_MIN,
                text_height_max=DEFAULT_TEXT_HEIGHT_MAX,
                max_match_distance=DEFAULT_MAX_MATCH_DISTANCE,
                text_reference_point="center",
            )
            return update_cached_result(filters)
        except Exception as e:
            last_raw_circles = saved_circles
            last_raw_texts = saved_texts
            last_raw_polylines = saved_polylines
            building_definitions = saved_buildings
            manual_overrides = saved_manual_overrides
            history_manual_overrides = saved_history_manual_overrides
            logger.exception("assign_buildings failed: %s", e)
            raise HTTPException(status_code=500, detail=str(e)) from e
    return apply_building_definitions(payload.buildings)


@app.post("/api/buildings/apply", response_model=CircleResponse)
async def apply_buildings(payload: BuildingApplyRequest) -> CircleResponse:
    return apply_building_definitions(payload.buildings)


@app.post("/api/manual-match", response_model=CircleResponse)
async def create_manual_match(payload: ManualMatchRequest) -> CircleResponse:
    if not last_raw_circles or not last_raw_texts:
        # 빈 결과 반환 (404 에러 방지)
        filters = current_filters or FilterSettings(
            min_diameter=DEFAULT_MIN_DIAMETER,
            max_diameter=DEFAULT_MAX_DIAMETER,
            text_height_min=DEFAULT_TEXT_HEIGHT_MIN,
            text_height_max=DEFAULT_TEXT_HEIGHT_MAX,
            max_match_distance=DEFAULT_MAX_MATCH_DISTANCE,
            text_reference_point="center",
        )
        return CircleResponse(
            summary=SummaryStats(total_circles=0, total_texts=0, matched_pairs=0, duplicate_groups=0),
            circles=[],
            texts=[],
            duplicates=[],
            errors=[],
            filter=filters,
            match_corrections=[],
        )
    circle_ids = {circle["id"] for circle in last_raw_circles}
    text_ids = {text["id"] for text in last_raw_texts}
    if payload.circle_id not in circle_ids:
        raise HTTPException(status_code=400, detail="Unknown circle id.")
    if payload.text_id not in text_ids:
        raise HTTPException(status_code=400, detail="Unknown text id.")
    manual_overrides[payload.circle_id] = payload.text_id
    filters = (
        payload.filter
        if payload.filter is not None
        else (current_filters or FilterSettings(
            min_diameter=DEFAULT_MIN_DIAMETER,
            max_diameter=DEFAULT_MAX_DIAMETER,
            text_height_min=DEFAULT_TEXT_HEIGHT_MIN,
            text_height_max=DEFAULT_TEXT_HEIGHT_MAX,
            max_match_distance=DEFAULT_MAX_MATCH_DISTANCE,
            text_reference_point="center",
        ))
    )
    return build_response_light(filters)


def _update_manual_history_from_saved_payload(data: Dict[str, Any]) -> None:
    update_manual_history(
        normalize_project_name(data.get("project")),
        normalize_source_type(data.get("sourceType")),
        circles=data.get("circles") or [],
        texts=data.get("texts") or [],
        manual_overrides=data.get("manualOverrides") or {},
        work_id=data.get("id"),
        work_title=data.get("title"),
        timestamp=data.get("timestamp"),
    )


@app.delete("/api/manual-match/{circle_id}", response_model=CircleResponse)
async def delete_manual_match(circle_id: str) -> CircleResponse:
    global history_manual_overrides
    manual_overrides.pop(circle_id, None)
    history_manual_overrides.pop(circle_id, None)
    filters = current_filters or FilterSettings(
        min_diameter=DEFAULT_MIN_DIAMETER,
        max_diameter=DEFAULT_MAX_DIAMETER,
        text_height_min=DEFAULT_TEXT_HEIGHT_MIN,
        text_height_max=DEFAULT_TEXT_HEIGHT_MAX,
        max_match_distance=DEFAULT_MAX_MATCH_DISTANCE,
        text_reference_point="center",
    )
    return build_response_light(filters)


# ---------- 중앙 저장: DXF 처리 결과 저장/불러오기 ----------


@app.get("/api/saved-works")
async def list_saved_works() -> List[Dict[str, Any]]:
    """저장된 작업 목록 반환 (id, title, timestamp, buildingCount, circleCount)."""
    index = _read_saved_works_index()
    index.sort(key=lambda x: (x.get("timestamp") or ""), reverse=True)
    return index


@app.post("/api/saved-works")
async def create_saved_work(payload: Dict[str, Any]) -> Dict[str, Any]:
    """처리 결과 JSON을 중앙에 저장. id는 서버에서 생성. title, author, sourceFileName, project 사용."""
    work_id = f"work_{uuid.uuid4().hex[:12]}"
    payload["id"] = work_id
    if "timestamp" not in payload or not payload["timestamp"]:
        from datetime import datetime, timezone
        payload["timestamp"] = datetime.now(timezone.utc).isoformat()
    title = (payload.get("title") or "").strip() or "이름 없음"
    payload["title"] = title
    payload.setdefault("author", (payload.get("author") or "").strip() or None)
    payload.setdefault("sourceFileName", payload.get("sourceFileName") or None)
    project = (payload.get("project") or "").strip() or "기본"
    payload["project"] = project
    payload["sourceType"] = normalize_source_type(payload.get("sourceType"))
    payload = apply_work_defaults(payload)
    _put_saved_work(work_id, payload)
    _update_manual_history_from_saved_payload(payload)
    index = _read_saved_works_index()
    entry = {
        "id": work_id,
        "title": title,
        "author": payload.get("author"),
        "sourceFileName": payload.get("sourceFileName"),
        "project": project,
        "sourceType": payload.get("sourceType"),
        "timestamp": payload.get("timestamp"),
        "buildingCount": len(payload.get("buildings") or []),
        "circleCount": len(payload.get("circles") or []),
    }
    index.insert(0, entry)
    _write_saved_works_index(index)
    return entry


@app.get("/api/saved-works/{work_id}")
async def get_saved_work(work_id: str) -> Dict[str, Any]:
    """저장된 작업 전체 JSON 반환."""
    data = _get_saved_work(work_id)
    if data is None:
        raise HTTPException(status_code=404, detail="Saved work not found.")
    try:
        return _rehydrate_saved_work_payload(data)
    except Exception as e:
        logger.exception("saved work rehydrate failed for %s: %s", work_id, e)
        return data


@app.put("/api/saved-works/{work_id}")
async def replace_saved_work(work_id: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    """저장된 작업 전체 덮어쓰기 (현재 화면 상태로 기존 작업 교체)."""
    if _get_saved_work(work_id) is None:
        raise HTTPException(status_code=404, detail="Saved work not found.")
    from datetime import datetime, timezone
    payload["id"] = work_id
    payload["timestamp"] = datetime.now(timezone.utc).isoformat()
    title = (payload.get("title") or "").strip() or "이름 없음"
    payload["title"] = title
    payload.setdefault("author", (payload.get("author") or "").strip() or None)
    payload.setdefault("sourceFileName", payload.get("sourceFileName") or None)
    project = (payload.get("project") or "").strip() or "기본"
    payload["project"] = project
    payload["sourceType"] = normalize_source_type(payload.get("sourceType"))
    payload = apply_work_defaults(payload)
    _put_saved_work(work_id, payload)
    _update_manual_history_from_saved_payload(payload)
    index = _read_saved_works_index()
    entry = {
        "id": work_id,
        "title": title,
        "author": payload.get("author"),
        "sourceFileName": payload.get("sourceFileName"),
        "project": project,
        "sourceType": payload.get("sourceType"),
        "timestamp": payload["timestamp"],
        "buildingCount": len(payload.get("buildings") or []),
        "circleCount": len(payload.get("circles") or []),
    }
    for i, e in enumerate(index):
        if e.get("id") == work_id:
            index[i] = entry
            break
    else:
        index.insert(0, entry)
    _write_saved_works_index(index)
    return entry


def _merge_entity_patches(
    entities: Any, patches: Any, *, id_key: str = "id"
) -> None:
    """저장된 circles/texts 배열에 id 기준으로 패치 병합 (수동 매칭 증분 저장용)."""
    if not isinstance(entities, list) or not patches:
        return
    if not isinstance(patches, list):
        return
    id_to_idx: Dict[str, int] = {}
    for i, item in enumerate(entities):
        if isinstance(item, dict) and item.get(id_key) is not None:
            id_to_idx[str(item[id_key])] = i
    for p in patches:
        if not isinstance(p, dict) or p.get(id_key) is None:
            continue
        sid = str(p[id_key])
        if sid in id_to_idx:
            entities[id_to_idx[sid]] = p


@app.patch("/api/saved-works/{work_id}")
async def update_saved_work(work_id: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    """저장된 작업의 제목·작성자·프로젝트 수정. 수동 매칭 후에는 manualOverrides·패치 필드로 증분 저장 가능."""
    from datetime import datetime, timezone

    data = _get_saved_work(work_id)
    if data is None:
        raise HTTPException(status_code=404, detail="Saved work not found.")
    touched = False
    if "title" in payload and payload["title"] is not None:
        title = (str(payload["title"]) or "").strip() or "이름 없음"
        data["title"] = title
        touched = True
    if "author" in payload:
        data["author"] = (str(payload["author"]) or "").strip() or None
        touched = True
    if "project" in payload and payload["project"] is not None:
        data["project"] = (str(payload["project"]) or "").strip() or "기본"
        touched = True

    if "sourceType" in payload and payload["sourceType"] is not None:
        data["sourceType"] = normalize_source_type(payload.get("sourceType"))
        touched = True

    manual_overrides_updated = False
    if "manualOverrides" in payload and payload["manualOverrides"] is not None:
        mo = payload["manualOverrides"]
        if not isinstance(mo, dict):
            raise HTTPException(status_code=400, detail="manualOverrides must be an object.")
        data["manualOverrides"] = {str(k): str(v) for k, v in mo.items()}
        touched = True
        manual_overrides_updated = True
    if "circlePatches" in payload and payload["circlePatches"] is not None:
        _merge_entity_patches(data.get("circles"), payload["circlePatches"])
        touched = True
    if "textPatches" in payload and payload["textPatches"] is not None:
        _merge_entity_patches(data.get("texts"), payload["textPatches"])
        touched = True
    if "summary" in payload and payload["summary"] is not None:
        data["summary"] = payload["summary"]
        touched = True
    if "errors" in payload and payload["errors"] is not None:
        data["errors"] = payload["errors"]
        touched = True
    if "duplicates" in payload and payload["duplicates"] is not None:
        data["duplicates"] = payload["duplicates"]
        touched = True

    # 수동 매칭 증분 저장: 에러·중복을 보내지 않으면 불러오기 시 클라이언트가 circles 기준으로 재계산
    if manual_overrides_updated and "errors" not in payload:
        data["errors"] = []
        data["duplicates"] = []
        touched = True

    if touched:
        data["timestamp"] = datetime.now(timezone.utc).isoformat()
    data = apply_work_defaults(data)
    _put_saved_work(work_id, data)
    _update_manual_history_from_saved_payload(data)
    index = _read_saved_works_index()
    for entry in index:
        if entry.get("id") == work_id:
            if "title" in data:
                entry["title"] = data["title"]
            if "author" in data:
                entry["author"] = data["author"]
            if "project" in data:
                entry["project"] = data["project"]
            if "sourceType" in data:
                entry["sourceType"] = data["sourceType"]
            if "timestamp" in data:
                entry["timestamp"] = data["timestamp"]
            if "circles" in data and isinstance(data["circles"], list):
                entry["circleCount"] = len(data["circles"])
            if "buildings" in data and isinstance(data["buildings"], list):
                entry["buildingCount"] = len(data["buildings"])
            break
    _write_saved_works_index(index)
    return {
        "id": work_id,
        "title": data.get("title"),
        "author": data.get("author"),
        "project": data.get("project"),
        "sourceType": data.get("sourceType"),
        "timestamp": data.get("timestamp"),
    }


@app.delete("/api/saved-works/{work_id}")
async def delete_saved_work(work_id: str) -> Dict[str, str]:
    """저장된 작업 삭제."""
    if not _get_saved_work(work_id):
        raise HTTPException(status_code=404, detail="Saved work not found.")
    _remove_saved_work(work_id)
    index = _read_saved_works_index()
    index[:] = [e for e in index if e.get("id") != work_id]
    _write_saved_works_index(index)
    return {"status": "deleted", "id": work_id}


# ---------- 설정(필터·동 외곽) 저장/불러오기 API ----------


@app.get("/api/saved-settings")
async def list_saved_settings(
    context_project: Optional[str] = Query(None, alias="contextProject"),
    context_project_id: Optional[str] = Query(None, alias="contextProjectId"),
) -> List[Dict[str, Any]]:
    """저장된 설정 목록 (projectName, versionId, timestamp, buildingCount)."""
    index = _read_saved_settings_index()
    context = _normalize_settings_context_project(context_project)
    context_id = _normalize_settings_context_project_id(context_project_id)
    if context_id:
        index = [entry for entry in index if _normalize_settings_context_project_id(entry.get("contextProjectId")) == context_id]
    if context:
        index = [entry for entry in index if _normalize_settings_context_project(entry.get("contextProject")) == context]
    index.sort(key=lambda x: (x.get("timestamp") or ""), reverse=True)
    return index


@app.post("/api/saved-settings")
async def create_saved_setting(payload: Dict[str, Any]) -> Dict[str, Any]:
    """설정 저장 (projectName, versionId, timestamp, buildings, clusterPolylines, filter, buildingCount, pendingNames)."""
    project_name = normalize_project_name(payload.get("projectName"))
    version_id = (payload.get("versionId") or "").strip()
    context_project = _normalize_settings_context_project(payload.get("contextProject"))
    context_project_id = _normalize_settings_context_project_id(payload.get("contextProjectId"))
    if not project_name or not version_id:
        raise HTTPException(status_code=400, detail="projectName and versionId are required.")
    payload = apply_setting_defaults(payload)
    payload["projectName"] = project_name
    if context_project:
        payload["contextProject"] = context_project
    if context_project_id:
        payload["contextProjectId"] = context_project_id
    _put_saved_setting(project_name, version_id, payload, context_project, context_project_id)
    index = _read_saved_settings_index()
    entry = {
        "name": project_name,
        "versionId": version_id,
        "timestamp": payload.get("timestamp", ""),
        "buildingCount": payload.get("buildingCount", 0),
        "sourceType": payload.get("sourceType"),
        "contextProject": context_project or "",
        "contextProjectId": context_project_id or "",
    }
    index = [
        e
        for e in index
        if not (
            e.get("name") == project_name
            and e.get("versionId") == version_id
            and _normalize_settings_context_project(e.get("contextProject")) == (context_project or "")
            and _normalize_settings_context_project_id(e.get("contextProjectId")) == (context_project_id or "")
        )
    ]
    index.insert(0, entry)
    _write_saved_settings_index(index)
    return {
        "projectName": project_name,
        "versionId": version_id,
        "sourceType": payload.get("sourceType"),
        "contextProject": context_project or "",
        "contextProjectId": context_project_id or "",
    }


@app.get("/api/saved-settings/item")
async def get_saved_setting(
    name: str = Query(..., alias="name"),
    version: str = Query(..., alias="version"),
    context_project: Optional[str] = Query(None, alias="contextProject"),
    context_project_id: Optional[str] = Query(None, alias="contextProjectId"),
) -> Dict[str, Any]:
    """설정 단건 조회."""
    context = _normalize_settings_context_project(context_project)
    context_id = _normalize_settings_context_project_id(context_project_id)
    data = _get_saved_setting(name, version, context, context_id)
    if data is None and (context or context_id):
        # 레거시 저장본(컨텍스트 없는 키) 호환
        data = _get_saved_setting(name, version, None, None)
    if data is None:
        raise HTTPException(status_code=404, detail="Saved setting not found.")
    return data


@app.delete("/api/saved-settings/item")
async def delete_saved_setting(
    name: str = Query(..., alias="name"),
    version: str = Query(..., alias="version"),
    context_project: Optional[str] = Query(None, alias="contextProject"),
    context_project_id: Optional[str] = Query(None, alias="contextProjectId"),
) -> Dict[str, str]:
    """설정 삭제."""
    context = _normalize_settings_context_project(context_project)
    context_id = _normalize_settings_context_project_id(context_project_id)
    if _get_saved_setting(name, version, context, context_id) is None:
        raise HTTPException(status_code=404, detail="Saved setting not found.")
    _remove_saved_setting(name, version, context, context_id)
    index = _read_saved_settings_index()
    index[:] = [
        e
        for e in index
        if not (
            e.get("name") == name
            and e.get("versionId") == version
            and _normalize_settings_context_project(e.get("contextProject")) == (context or "")
            and _normalize_settings_context_project_id(e.get("contextProjectId")) == (context_id or "")
        )
    ]
    _write_saved_settings_index(index)
    return {"status": "deleted"}


def _parse_json_list_field(raw: str, field_name: str) -> List[Dict[str, Any]]:
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail=f"{field_name} is not valid JSON.") from exc
    if not isinstance(data, list):
        raise HTTPException(status_code=400, detail=f"{field_name} must be a JSON list.")
    return data


async def _parse_json_list_upload_field(
    raw: Optional[str],
    upload: Optional[UploadFile],
    field_name: str,
) -> List[Dict[str, Any]]:
    payload = raw
    if upload is not None:
        content = await upload.read()
        if not content:
            raise HTTPException(status_code=400, detail=f"{field_name} file is empty.")
        try:
            payload = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise HTTPException(status_code=400, detail=f"{field_name} file must be UTF-8 JSON.") from exc
    if payload is None:
        raise HTTPException(status_code=400, detail=f"{field_name} is required.")
    return _parse_json_list_field(payload, field_name)


@app.post("/api/excel/inspect")
async def inspect_excel(file: UploadFile = File(...)) -> Dict[str, Any]:
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Excel file is empty.")
    try:
        return inspect_excel_workbook(content)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to inspect Excel file: {exc}") from exc


@app.post("/api/excel/compare")
async def compare_excel_coordinates(
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Form(None),
    sheet_names_json: Optional[str] = Form(None),
    header_row: int = Form(...),
    number_column: str = Form(...),
    x_column: str = Form(...),
    y_column: str = Form(...),
    version_a_circles_json: Optional[str] = Form(None),
    version_b_circles_json: Optional[str] = Form(None),
    version_a_circles_file: Optional[UploadFile] = File(None),
    version_b_circles_file: Optional[UploadFile] = File(None),
    building_column: Optional[str] = Form(None),
    building_source_mode: Optional[str] = Form(None),
    use_sheet_name_as_building: Optional[bool] = Form(True),
    coord_tolerance: float = Form(0.01),
    version_a_label: Optional[str] = Form(None),
    version_b_label: Optional[str] = Form(None),
) -> Dict[str, Any]:
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Excel file is empty.")
    circles_a = await _parse_json_list_upload_field(
        version_a_circles_json,
        version_a_circles_file,
        "version_a_circles_json",
    )
    circles_b = await _parse_json_list_upload_field(
        version_b_circles_json,
        version_b_circles_file,
        "version_b_circles_json",
    )
    sheet_names = (
        [str(item) for item in _parse_json_list_field(sheet_names_json, "sheet_names_json")]
        if sheet_names_json
        else ([sheet_name] if sheet_name else [])
    )
    if not sheet_names:
        raise HTTPException(status_code=400, detail="At least one sheet must be selected.")
    try:
        result = compare_excel_workbook(
            content,
            sheet_name=sheet_name,
            sheet_names=sheet_names,
            header_row=header_row,
            building_column=building_column,
            number_column=number_column,
            x_column=x_column,
            y_column=y_column,
            building_source_mode=building_source_mode,
            use_sheet_name_as_building=use_sheet_name_as_building,
            circles_a=circles_a,
            circles_b=circles_b,
            coord_tolerance=coord_tolerance,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to compare Excel file: {exc}") from exc
    result["versionLabels"] = {
        "old": version_a_label or "이전 버전",
        "new": version_b_label or "현재 버전",
    }
    return result


@app.get("/api/construction/datasets")
async def list_construction_dataset_items(
    project_context: Optional[str] = Query(None, alias="project_context"),
) -> List[Dict[str, Any]]:
    return list_construction_datasets(project_context=project_context)


@app.delete("/api/construction/datasets/{dataset_id}")
async def delete_construction_dataset_item(dataset_id: str) -> Dict[str, Any]:
    try:
        return delete_construction_dataset(dataset_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except Exception as exc:
        logger.exception("Construction dataset delete failed: %s", exc)
        raise HTTPException(status_code=500, detail="Failed to delete construction dataset.") from exc


@app.post("/api/construction/import-workbook")
async def import_construction_workbook(
    file: UploadFile = File(...),
    dataset_name: Optional[str] = Form(None),
    source_url: Optional[str] = Form(None),
    report_id: Optional[int] = Form(None),
    project_context: Optional[str] = Form(None),
) -> Dict[str, Any]:
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Construction workbook is empty.")
    try:
        return import_construction_workbook_bytes(
            content,
            filename=file.filename,
            source_type="manual-upload",
            source_url=source_url,
            report_id=report_id,
            dataset_name=dataset_name,
            project_context=project_context,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        logger.exception("Construction workbook import failed: %s", exc)
        raise HTTPException(status_code=500, detail="Failed to import construction workbook.") from exc


@app.post("/api/construction/sync-pdam")
async def sync_construction_pdam(payload: ConstructionSyncRequest) -> Dict[str, Any]:
    try:
        return sync_pdam_workbook(
            user_id=payload.user_id,
            password=payload.password,
            report_page_url=payload.report_page_url,
            report_id=payload.report_id,
            source_url=payload.source_url or "https://we8104.com/",
            dataset_name=payload.dataset_name,
            project_context=payload.project_context,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        logger.exception("PDAM sync failed: %s", exc)
        raise HTTPException(status_code=500, detail="PDAM sync failed.") from exc


@app.post("/api/construction/dashboard")
async def get_construction_dashboard(payload: ConstructionDashboardRequest) -> Dict[str, Any]:
    try:
        circles = payload.circles
        if not circles and payload.work_id:
            saved = _get_saved_work(payload.work_id)
            if saved and isinstance(saved.get("circles"), list):
                circles = saved["circles"]
        return build_construction_dashboard(
            payload.dataset_id,
            circles=circles,
            work_id=payload.work_id,
            date_from=payload.date_from,
            date_to=payload.date_to,
            month=payload.month,
            equipment=payload.equipment,
            method=payload.method,
            location=payload.location,
            equipments=payload.equipments,
            methods=payload.methods,
            locations=payload.locations,
            remaining_threshold=payload.remaining_threshold,
            settlement_month=payload.settlement_month,
            settlement_start_day=payload.settlement_start_day,
            settlement_end_day=payload.settlement_end_day,
            exclude_identical_geometry_duplicates=payload.exclude_identical_geometry_duplicates,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        logger.exception("Construction dashboard failed: %s", exc)
        raise HTTPException(status_code=500, detail="Failed to build construction dashboard.") from exc


def _meissa_access_token(authorization: Optional[str]) -> str:
    if not authorization or not str(authorization).strip():
        raise HTTPException(status_code=401, detail="Authorization 헤더에 Meissa 액세스 토큰이 필요합니다. (예: JWT …)")
    auth = str(authorization).strip()
    low = auth[:7].lower()
    if low == "bearer ":
        auth = auth[7:].strip()
    elif auth[:4].upper() == "JWT ":
        auth = auth[4:].strip()
    if not auth:
        raise HTTPException(status_code=401, detail="유효하지 않은 Authorization 헤더입니다.")
    return auth


def _meissa_access_token_header_or_query(
    authorization: Optional[str],
    access_token: Optional[str],
) -> str:
    """헤더 우선. <img src> 는 Authorization 을 못 쓰므로 access_token 쿼리 허용(서버 로그에 노출될 수 있음)."""
    if authorization and str(authorization).strip():
        return _meissa_access_token(authorization)
    raw = (access_token or "").strip()
    if not raw:
        raise HTTPException(
            status_code=401,
            detail="Authorization 헤더 또는 쿼리 access_token 이 필요합니다.",
        )
    return _meissa_access_token(raw if raw.lower().startswith("jwt ") else f"JWT {raw}")


@app.post("/api/meissa/login")
async def api_meissa_login(payload: MeissaLoginRequest) -> Dict[str, Any]:
    try:
        if (payload.verification_code or "").strip():
            out = await run_in_threadpool(
                meissa_login_with_verification,
                email=payload.email or "",
                password=payload.password or "",
                verification_code=payload.verification_code.strip(),
                service=payload.service or "cloud",
            )
        else:
            out = await run_in_threadpool(
                meissa_login,
                email=payload.email or "",
                password=payload.password or "",
                service=payload.service or "cloud",
            )
        out.pop("raw", None)
        return out
    except Meissa2FARequired:
        return {"needsVerification": True}
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        logger.exception("Meissa login error: %s", exc)
        raise HTTPException(status_code=500, detail="Meissa 로그인 처리 중 오류가 발생했습니다.") from exc


@app.get("/api/meissa/projects")
async def api_meissa_projects(authorization: Optional[str] = Header(None)) -> Dict[str, Any]:
    token = _meissa_access_token(authorization)
    try:
        return {"projects": meissa_list_projects(token)}
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/meissa/projects/{project_id}/zones")
async def api_meissa_zones(project_id: str, authorization: Optional[str] = Header(None)) -> Dict[str, Any]:
    token = _meissa_access_token(authorization)
    try:
        return {"zones": meissa_list_zones(token, project_id)}
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/meissa/zones/{zone_id}/snapshots")
async def api_meissa_snapshots(zone_id: str, authorization: Optional[str] = Header(None)) -> Dict[str, Any]:
    token = _meissa_access_token(authorization)
    try:
        return {"snapshots": meissa_list_snapshots(token, zone_id)}
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/meissa/snapshots/{snapshot_id}/resources")
async def api_meissa_snapshot_resources(snapshot_id: str, authorization: Optional[str] = Header(None)) -> Dict[str, Any]:
    token = _meissa_access_token(authorization)
    try:
        return {"resources": meissa_list_snapshot_resources(token, snapshot_id)}
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/meissa/snapshots/{snapshot_id}/resources/{resource_id}/points")
async def api_meissa_resource_points_sample(
    snapshot_id: str,
    resource_id: str,
    limit: int = Query(8000, ge=100, le=30000),
    phase: int = Query(0, ge=0, le=100000),
    authorization: Optional[str] = Header(None),
) -> Dict[str, Any]:
    token = _meissa_access_token(authorization)
    try:
        return meissa_sample_resource_points(token, snapshot_id, resource_id, limit=limit, phase=phase)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/meissa/snapshots/{snapshot_id}/nearest-z")
async def api_meissa_nearest_z(
    snapshot_id: str,
    x: float = Query(..., description="포인트클라우드와 동일 CRS 평면 X"),
    y: float = Query(..., description="포인트클라우드와 동일 CRS 평면 Y"),
    resource_id: Optional[str] = Query(None, description="지정 시 해당 리소스만 샘플"),
    limit: int = Query(8000, ge=500, le=30000),
    max_phases: int = Query(4, ge=1, le=8),
    zone_id: Optional[str] = Query(
        None,
        description="Meissa 존 ID(선택). 생략 시 스냅샷 상세 API로 존 추론 후에도 auto 동작.",
    ),
    z_source: Optional[str] = Query(
        None,
        description="생략 시 auto. auto=DSM 포인트 어노테이션 우선·실패 시 점군, pointcloud=점군만, cloud=DSM만.",
    ),
    authorization: Optional[str] = Header(None),
) -> Dict[str, Any]:
    token = _meissa_access_token(authorization)
    zs = (z_source or "").strip().lower()
    if not zs:
        # zone_id 없어도 스냅샷 상세에서 존 추론 → DSM 포인트 어노테이션 우선(실패 시 점군). URL에 zone 생략 시에도 웹과 가깝게.
        zs = "auto"
    try:
        return await run_in_threadpool(
            meissa_nearest_z_xy_combined,
            token,
            snapshot_id,
            x,
            y,
            zone_id=zone_id,
            z_source=zs,
            resource_id=resource_id,
            limit=limit,
            max_phases=max_phases,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/meissa/snapshots/{snapshot_id}/dsm-z-batch")
async def api_meissa_dsm_z_batch(
    snapshot_id: str,
    project_id: str = Query(..., min_length=1),
    body: Dict[str, Any] = Body(...),
    authorization: Optional[str] = Header(None),
) -> Dict[str, Any]:
    """
    Carta export/dsm/dsm.tif 1회 수신 후 (x,y) 목록에 대해 고도 Z 일괄 샘플.
    비교 계산에서 nearest-z 수천 회 호출 대신 사용.
    """
    token = _meissa_access_token(authorization)
    pts = body.get("points")
    if not isinstance(pts, list):
        raise HTTPException(status_code=400, detail="JSON body에 points 배열이 필요합니다.")
    return await run_in_threadpool(
        meissa_dsm_z_batch_from_carta_export,
        token,
        project_id,
        snapshot_id,
        pts,
    )


@app.get("/api/meissa/snapshots/{snapshot_id}/overlay-2d-image")
async def api_meissa_overlay_2d_image(snapshot_id: str, authorization: Optional[str] = Header(None)) -> Dict[str, Any]:
    token = _meissa_access_token(authorization)
    try:
        return meissa_get_snapshot_overlay_2d_image(token, snapshot_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/meissa/snapshots/{snapshot_id}/overlay-2d-image/raw")
async def api_meissa_overlay_2d_image_raw(snapshot_id: str, authorization: Optional[str] = Header(None)):
    token = _meissa_access_token(authorization)
    try:
        data = meissa_get_snapshot_overlay_2d_binary(token, snapshot_id)
        if not data.get("ok"):
            raise HTTPException(
                status_code=404,
                detail={
                    "message": str(data.get("message") or "2D 이미지 없음"),
                    "snapshotId": data.get("snapshotId") or snapshot_id,
                    "debug": data.get("debug"),
                },
            )
        body = data.get("body")
        if not isinstance(body, (bytes, bytearray)) or not body:
            raise HTTPException(
                status_code=404,
                detail={
                    "message": "2D 이미지 바이트가 비어 있습니다.",
                    "snapshotId": data.get("snapshotId") or snapshot_id,
                    "debug": data.get("debug"),
                },
            )
        ctype = str(data.get("contentType") or "image/png")
        # StreamingResponse(BytesIO)는 동시 요청·Win/py3.12 asyncio에서 빈 write 콜백 assert 유발 사례 있음
        return Response(content=bytes(body), media_type=ctype)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/meissa/snapshots/{snapshot_id}/overlay-2d-georef")
async def api_meissa_overlay_2d_georef(snapshot_id: str, authorization: Optional[str] = Header(None)) -> Dict[str, Any]:
    token = _meissa_access_token(authorization)
    try:
        return meissa_get_snapshot_overlay_2d_georef(token, snapshot_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


def _file_response_png_if_nonempty(path: Optional[str], headers: Dict[str, str]) -> Optional[FileResponse]:
    """0바이트 PNG 캐시는 FileResponse 가 빈 소켓 쓰기로 asyncio assert 를 유발할 수 있어 제외."""
    if not path or not isinstance(path, str):
        return None
    try:
        if os.path.isfile(path) and os.path.getsize(path) > 0:
            return FileResponse(path, media_type="image/png", headers=headers)
    except OSError:
        pass
    return None


@app.get("/api/meissa/snapshots/{snapshot_id}/orthophoto-preview")
async def api_meissa_orthophoto_preview(
    background_tasks: BackgroundTasks,
    snapshot_id: str,
    project_id: str = Query(..., min_length=1),
    max_edge: Optional[int] = Query(None, ge=1024, le=16384),
    crop_x: Optional[int] = Query(None, ge=0),
    crop_y: Optional[int] = Query(None, ge=0),
    crop_w: Optional[int] = Query(None, ge=1, le=8192),
    crop_h: Optional[int] = Query(None, ge=1, le=8192),
    access_token: Optional[str] = Query(
        None,
        description="선택. <img src> 로 정사를 직접 불러올 때(헤더 불가). 값은 액세스 JWT.",
    ),
    authorization: Optional[str] = Header(None),
):
    """
    기본: Carta 최대 export PNG 를 바이트 그대로(또는 디스크 full 캐시) 제공.
    Pillow 축소 결과는 edge 캐시. TIF 폴백은 백엔드 환경변수로만 활성화.
    crop_x/y/w/h 가 모두 있으면 **디스크 full export** 기준 픽셀 직사각형만 잘라 max_edge 이하로 인코딩(줌 영역 고해상).
    """
    token = _meissa_access_token_header_or_query(authorization, access_token)
    edge_eff = meissa_orthophoto_effective_preview_edge(max_edge)
    if (
        crop_x is not None
        and crop_y is not None
        and crop_w is not None
        and crop_h is not None
    ):
        try:
            data = await run_in_threadpool(
                meissa_orthophoto_full_export_crop_to_png_bytes,
                project_id,
                snapshot_id,
                int(crop_x),
                int(crop_y),
                int(crop_w),
                int(crop_h),
                int(edge_eff),
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        if not data.get("ok"):
            reason = str(data.get("reason") or "")
            if reason == "nocache":
                raise HTTPException(
                    status_code=503,
                    detail={
                        "message": str(data.get("message") or "full ortho cache not ready"),
                        "reason": "nocache",
                        "snapshotId": data.get("snapshotId") or snapshot_id,
                        "projectId": data.get("projectId") or project_id,
                    },
                )
            raise HTTPException(
                status_code=404,
                detail={
                    "message": str(data.get("message") or "orthophoto crop 실패"),
                    "snapshotId": data.get("snapshotId") or snapshot_id,
                    "projectId": data.get("projectId") or project_id,
                },
            )
        body = data.get("body")
        if not isinstance(body, (bytes, bytearray)) or not body:
            raise HTTPException(
                status_code=404,
                detail={"message": "crop PNG 바이트가 비어 있습니다.", "snapshotId": snapshot_id, "projectId": project_id},
            )
        body_bytes = bytes(body)
        headers = {
            "Cache-Control": "private, max-age=120",
            "X-Ortho-Source": "full-export-crop",
            "X-Ortho-Max-Edge": str(int(edge_eff)),
        }
        for hk, hv in (
            ("X-Ortho-Full-W", data.get("full_px_w")),
            ("X-Ortho-Full-H", data.get("full_px_h")),
            ("X-Ortho-Src-X0", data.get("src_x0")),
            ("X-Ortho-Src-Y0", data.get("src_y0")),
            ("X-Ortho-Src-W", data.get("src_w")),
            ("X-Ortho-Src-H", data.get("src_h")),
            ("X-Ortho-Out-W", data.get("width")),
            ("X-Ortho-Out-H", data.get("height")),
        ):
            try:
                if hv is not None and str(hv).strip() != "":
                    headers[hk] = str(int(hv))
            except (TypeError, ValueError):
                pass
        return Response(content=body_bytes, media_type="image/png", headers=headers)
    # full-export 디스크 캐시는 "최대 해상도" 요청에만 맞춘다. 그렇지 않으면 max_edge=6144 도
    # 동일한 원본 PNG가 나가 프론트의 저해상→고해상 2단계가 같은 픽셀이 되어 화면이 안 바뀐다.
    _ortho_full_edge = 16384
    try:
        full_cached = await run_in_threadpool(
            meissa_orthophoto_disk_cache_full_export_path_if_valid, project_id, snapshot_id
        )
        if full_cached and int(edge_eff) >= _ortho_full_edge:
            fr = _file_response_png_if_nonempty(
                full_cached,
                {
                    "X-Ortho-Source": "disk-cache-full",
                    "X-Ortho-Full-Export": "1",
                    "X-Ortho-Max-Edge": str(int(edge_eff)),
                    "Cache-Control": "public, max-age=604800",
                },
            )
            if fr is not None:
                return fr
    except Exception:
        pass
    try:
        cached_path = await run_in_threadpool(
            meissa_orthophoto_disk_cache_path_if_valid, project_id, snapshot_id, edge_eff
        )
        if cached_path:
            fr = _file_response_png_if_nonempty(
                cached_path,
                {
                    "X-Ortho-Source": "disk-cache",
                    "X-Ortho-Max-Edge": str(int(edge_eff)),
                    "Cache-Control": "public, max-age=604800",
                },
            )
            if fr is not None:
                return fr
    except Exception:
        pass
    try:
        data = await run_in_threadpool(
            meissa_get_carta_orthophoto_preview_png,
            token,
            project_id,
            snapshot_id,
            max_edge=edge_eff,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if not data.get("ok"):
        fb = await run_in_threadpool(
            meissa_orthophoto_disk_cache_best_valid_path_up_to,
            project_id,
            snapshot_id,
            int(edge_eff),
        )
        if fb:
            pth, el_fb = fb
            hdrs: Dict[str, str] = {
                "Cache-Control": "public, max-age=3600",
                "X-Ortho-Source": "disk-cache-fallback",
                "X-Ortho-Max-Edge": str(int(el_fb)),
            }
            if int(el_fb) < int(edge_eff):
                hdrs["X-Ortho-Requested-Max-Edge"] = str(int(edge_eff))
            fr_fb = _file_response_png_if_nonempty(pth, hdrs)
            if fr_fb is not None:
                return fr_fb
        raise HTTPException(
            status_code=404,
            detail={
                "message": str(data.get("message") or "orthophoto 미리보기 실패"),
                "snapshotId": data.get("snapshotId") or snapshot_id,
                "projectId": data.get("projectId") or project_id,
            },
        )
    body = data.get("body")
    if not isinstance(body, (bytes, bytearray)) or not body:
        fb = await run_in_threadpool(
            meissa_orthophoto_disk_cache_best_valid_path_up_to,
            project_id,
            snapshot_id,
            int(edge_eff),
        )
        if fb:
            pth, el_fb = fb
            hdrs2: Dict[str, str] = {
                "Cache-Control": "public, max-age=3600",
                "X-Ortho-Source": "disk-cache-fallback",
                "X-Ortho-Max-Edge": str(int(el_fb)),
            }
            if int(el_fb) < int(edge_eff):
                hdrs2["X-Ortho-Requested-Max-Edge"] = str(int(edge_eff))
            fr_fb2 = _file_response_png_if_nonempty(pth, hdrs2)
            if fr_fb2 is not None:
                return fr_fb2
        raise HTTPException(
            status_code=404,
            detail={"message": "PNG 바이트가 비어 있습니다.", "snapshotId": snapshot_id, "projectId": project_id},
        )
    headers = {"Cache-Control": "public, max-age=604800"}
    src = data.get("source")
    if isinstance(src, str) and src.strip():
        headers["X-Ortho-Source"] = src.strip()[:120]
    body_bytes = bytes(body)
    if data.get("disk_cache_slot") == "full_export":
        headers["X-Ortho-Full-Export"] = "1"
        w0 = int(data.get("width") or 0)
        h0 = int(data.get("height") or 0)
        headers["X-Ortho-Max-Edge"] = str(max(w0, h0) if (w0 or h0) else int(edge_eff))
        background_tasks.add_task(
            meissa_orthophoto_write_disk_cache_full_export, project_id, snapshot_id, body_bytes
        )
    else:
        headers["X-Ortho-Max-Edge"] = str(int(edge_eff))
        background_tasks.add_task(meissa_orthophoto_write_disk_cache, project_id, snapshot_id, body_bytes, edge_eff)
    return Response(content=body_bytes, media_type="image/png", headers=headers)


@app.get("/api/circles/export")
async def export_circles(
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    min_diameter: Optional[float] = Query(None),
    max_diameter: Optional[float] = Query(None),
    min_area: Optional[float] = Query(None),
    max_area: Optional[float] = Query(None),
    text_height_min: Optional[float] = Query(None),
    text_height_max: Optional[float] = Query(None),
    max_match_distance: Optional[float] = Query(None),
    text_reference_point: Optional[str] = Query(None),
    expected_buildings: Optional[int] = Query(None),
):
    global last_result
    filters_override: Optional[FilterSettings] = None
    if any(
        value is not None
        for value in (min_diameter, max_diameter, min_area, max_area, text_height_min, text_height_max, max_match_distance, text_reference_point)
    ):
        filters_override = resolve_filter_values(
            min_diameter, max_diameter, min_area, max_area,
            text_height_min, text_height_max, max_match_distance, text_reference_point
        )
    if expected_buildings is not None:
        set_expected_building_clusters(expected_buildings)
        if filters_override is None:
            filters_override = current_filters or FilterSettings(
                min_diameter=DEFAULT_MIN_DIAMETER,
                max_diameter=DEFAULT_MAX_DIAMETER,
                text_height_min=DEFAULT_TEXT_HEIGHT_MIN,
                text_height_max=DEFAULT_TEXT_HEIGHT_MAX,
                max_match_distance=DEFAULT_MAX_MATCH_DISTANCE,
                text_reference_point="center",
            )
    if filters_override is not None:
        last_result = update_cached_result(filters_override)
    # 데이터가 없어도 빈 결과로 반환 (404 에러 방지)
    if not last_result:
        last_result = CircleResponse(
            summary=SummaryStats(total_circles=0, total_texts=0, matched_pairs=0, duplicate_groups=0),
            circles=[],
            texts=[],
            duplicates=[],
            errors=[],
            filter=filters_override or FilterSettings(
                min_diameter=DEFAULT_MIN_DIAMETER,
                max_diameter=DEFAULT_MAX_DIAMETER,
                text_height_min=DEFAULT_TEXT_HEIGHT_MIN,
                text_height_max=DEFAULT_TEXT_HEIGHT_MAX,
                max_match_distance=DEFAULT_MAX_MATCH_DISTANCE,
                text_reference_point="center",
            ),
            match_corrections=[],
        )
    if not last_result.circles:
        # 빈 결과로 빈 파일 반환
        rows, error_rows, summary_rows, building_order = [], [], [], []
    else:
        rows, error_rows, summary_rows, building_order = prepare_export_rows(
            last_result.circles, last_result.texts, last_result.errors, building_definitions
        )
    return _build_export_stream(rows, error_rows, summary_rows, building_order, format)


@app.post("/api/circles/export")
async def export_circles_from_client(
    body: ExportFromClientBody = Body(...),
    format: str = Query("xlsx", pattern="^(csv|xlsx)$"),
):
    """
    동 윤곽선 적용된 현재 화면 데이터(circles/texts/errors/buildings)로 XLSX/CSV 다운로드.
    클라이언트가 보낸 circles 기준으로 building_name 등이 반영된 좌표로 내보냅니다.
    """
    if not body.circles or len(body.circles) == 0:
        raise HTTPException(
            status_code=400,
            detail="circles is required and must not be empty for client export.",
        )
    circle_records = [to_circle_model(c) for c in body.circles]
    text_list = body.texts or []
    text_records = [to_text_model(t) for t in text_list]
    error_list: List[MatchError] = []
    for e in body.errors or []:
        if isinstance(e, dict):
            try:
                error_list.append(MatchError(**e))
            except Exception:
                pass
        elif isinstance(e, MatchError):
            error_list.append(e)
    building_defs: List[BuildingDefinition] = []
    for b in body.buildings or []:
        if isinstance(b, dict):
            name = b.get("name") or "Unassigned"
            verts = []
            for v in b.get("vertices") or []:
                if isinstance(v, dict) and "x" in v and "y" in v:
                    verts.append(BuildingVertex(x=float(v["x"]), y=float(v["y"])))
            if not verts:
                verts = [BuildingVertex(x=0.0, y=0.0)]
            raw_kind = str(b.get("kind") or "").strip().lower()
            if raw_kind == "parking":
                bd_kind = "parking"
            elif raw_kind in ("tower_crane", "tower-crane", "tower"):
                bd_kind = "tower_crane"
            else:
                bd_kind = "building"
            building_defs.append(
                BuildingDefinition(
                    name=name,
                    kind=bd_kind,
                    slot=int(b["slot"]) if isinstance(b, dict) and b.get("slot") is not None else None,
                    vertices=verts,
                )
            )
    rows, error_rows, summary_rows, building_order = prepare_export_rows(
        circle_records, text_records, error_list, building_defs
    )
    return _build_export_stream(rows, error_rows, summary_rows, building_order, format)
